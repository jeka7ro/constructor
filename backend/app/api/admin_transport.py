"""
Admin API — Modul Transport: Foi de Parcurs
Gestionare drumuri vehicule cu GPS, calcul automat km / viteză / durată.
Validare program de lucru: nu se permit drumuri în afara orelor/zilelor configurate.
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, desc
from typing import List, Optional
from pydantic import BaseModel, Field
from datetime import datetime, date, time as dt_time
import uuid
import io
import math

from app.database import get_db
from app.models import TripLog, TripGPSPoint, Vehicle, User, ConstructionSite, Admin, Organization
from app.api.admin_auth import get_current_admin

router = APIRouter(prefix="/admin/transport", tags=["admin-transport"])

# =============================================================================
# PYDANTIC SCHEMAS
# =============================================================================

class GPSPointIn(BaseModel):
    latitude: float
    longitude: float
    speed_kmh: Optional[float] = None
    accuracy_m: Optional[float] = None
    altitude_m: Optional[float] = None
    timestamp: str  # ISO 8601

class TripCreate(BaseModel):
    vehicle_id: str
    driver_id: str
    site_id: Optional[str] = None
    date: Optional[str] = None           # YYYY-MM-DD, default today
    purpose_category: Optional[str] = "transport_materiale"
    purpose_notes: Optional[str] = None
    start_odometer: Optional[float] = None
    start_address: Optional[str] = None
    start_lat: Optional[float] = None
    start_lng: Optional[float] = None
    notes: Optional[str] = None

class TripEnd(BaseModel):
    end_odometer: Optional[float] = None
    end_address: Optional[str] = None
    end_lat: Optional[float] = None
    end_lng: Optional[float] = None
    notes: Optional[str] = None

class TripUpdate(BaseModel):
    vehicle_id: Optional[str] = None
    driver_id: Optional[str] = None
    site_id: Optional[str] = None
    purpose_category: Optional[str] = None
    purpose_notes: Optional[str] = None
    start_odometer: Optional[float] = None
    end_odometer: Optional[float] = None
    start_address: Optional[str] = None
    end_address: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None

class ApprovalPayload(BaseModel):
    action: str   # 'approve' | 'reject'
    rejection_note: Optional[str] = None

class TransportScheduleUpdate(BaseModel):
    transport_start_time: Optional[str] = None        # "HH:MM"
    transport_end_time: Optional[str] = None          # "HH:MM"
    transport_allowed_days: Optional[List[int]] = None # [0..6], 0=Luni
    transport_strict_schedule: Optional[bool] = None

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

DAY_NAMES_RO = ["Luni", "Marți", "Miercuri", "Joi", "Vineri", "Sâmbătă", "Duminică"]
EARTH_RADIUS_KM = 6371.0

def _parse_hhmm(s: str) -> dt_time:
    """Parse 'HH:MM' to time object."""
    h, m = map(int, s.split(":"))
    return dt_time(h, m)

def _check_schedule(org: Organization, now_local: datetime) -> dict:
    """
    Verifică dacă momentul curent este în fereastra de program transport.
    Returnează: { 'allowed': bool, 'out_of_schedule': bool, 'note': str }
    """
    if not org:
        return {"allowed": True, "out_of_schedule": False, "note": None}

    issues = []

    # Ziua săptămânii: weekday() = 0 (Luni) ... 6 (Duminică)
    allowed_days = org.transport_allowed_days or [0,1,2,3,4]
    if isinstance(allowed_days, str):
        import json
        try:
            allowed_days = json.loads(allowed_days)
        except:
            allowed_days = [int(x.strip()) for x in allowed_days.split(",") if x.strip().isdigit()]
    weekday = now_local.weekday()
    if weekday not in allowed_days:
        day_name = DAY_NAMES_RO[weekday]
        allowed_names = ", ".join(DAY_NAMES_RO[d] for d in sorted(allowed_days))
        issues.append(f"{day_name} nu este o zi lucrătoare de transport (permise: {allowed_names}).")

    # Intervalul orar
    start_str = org.transport_start_time or "06:00"
    end_str   = org.transport_end_time   or "20:00"
    t_start = _parse_hhmm(start_str)
    t_end   = _parse_hhmm(end_str)
    t_now   = now_local.time().replace(second=0, microsecond=0)

    if not (t_start <= t_now <= t_end):
        issues.append(
            f"Ora curentă {t_now.strftime('%H:%M')} este în afara programului "
            f"de transport ({start_str}–{end_str})."
        )

    out = bool(issues)
    note = " ".join(issues) if issues else None
    strict = org.transport_strict_schedule or False

    return {
        "allowed": not out or not strict,   # blocat doar dacă strict=True
        "out_of_schedule": out,
        "note": note,
        "start_time": start_str,
        "end_time": end_str,
        "allowed_days": allowed_days,
        "strict": strict,
    }

def haversine_km(lat1, lng1, lat2, lng2) -> float:
    """Calculate distance in km between two GPS coordinates."""
    r = EARTH_RADIUS_KM
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lng2 - lng1)
    a = math.sin(dphi/2)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda/2)**2
    return 2 * r * math.asin(math.sqrt(a))


def compute_trip_stats(trip: TripLog) -> dict:
    """Compute km, duration, avg speed and visited points from a TripLog."""
    gps = trip.gps_points  # already ordered by timestamp

    # ── Distance ──────────────────────────────────────────────────────────────
    distance_km = trip.distance_km  # use odometer value if set
    if distance_km is None and len(gps) >= 2:
        distance_km = sum(
            haversine_km(gps[i].latitude, gps[i].longitude,
                         gps[i+1].latitude, gps[i+1].longitude)
            for i in range(len(gps) - 1)
        )
        distance_km = round(distance_km, 2)

    # ── Duration ──────────────────────────────────────────────────────────────
    duration_minutes = None
    if trip.start_time and trip.end_time:
        duration_minutes = round((trip.end_time - trip.start_time).total_seconds() / 60, 1)
    elif trip.start_time and gps:
        last_ts = gps[-1].timestamp
        duration_minutes = round((last_ts - trip.start_time).total_seconds() / 60, 1)

    # ── Average speed ─────────────────────────────────────────────────────────
    avg_speed_kmh = None
    speeds = [p.speed_kmh for p in gps if p.speed_kmh is not None]
    if speeds:
        avg_speed_kmh = round(sum(speeds) / len(speeds), 1)
    elif distance_km and duration_minutes and duration_minutes > 0:
        avg_speed_kmh = round((distance_km / (duration_minutes / 60)), 1)

    max_speed_kmh = round(max(speeds), 1) if speeds else None

    # ── Visited GPS points (downsampled for response — max 200 pts) ───────────
    sampled_points = gps
    if len(gps) > 200:
        step = len(gps) // 200
        sampled_points = gps[::step]

    visited_points = [
        {
            "lat": p.latitude,
            "lng": p.longitude,
            "speed_kmh": p.speed_kmh,
            "timestamp": p.timestamp.isoformat() if p.timestamp else None,
        }
        for p in sampled_points
    ]

    return {
        "distance_km": distance_km,
        "duration_minutes": duration_minutes,
        "avg_speed_kmh": avg_speed_kmh,
        "max_speed_kmh": max_speed_kmh,
        "gps_points_count": len(gps),
        "visited_points": visited_points,
    }


def trip_to_dict(trip: TripLog, include_gps: bool = False) -> dict:
    """Serialize a TripLog to dict with computed stats."""
    stats = compute_trip_stats(trip)

    return {
        "id": trip.id,
        "organization_id": trip.organization_id,
        # Vehicle
        "vehicle_id": trip.vehicle_id,
        "vehicle_name": trip.vehicle.name if trip.vehicle else None,
        "vehicle_plate": trip.vehicle.plate_number if trip.vehicle else None,
        # Driver
        "driver_id": trip.driver_id,
        "driver_name": trip.driver.full_name if trip.driver else None,
        "driver_code": trip.driver.employee_code if trip.driver else None,
        # Site destination
        "site_id": trip.site_id,
        "site_name": trip.site.name if trip.site else None,
        # Dates & time
        "date": trip.date.isoformat() if trip.date else None,
        "status": trip.status,
        "start_time": trip.start_time.isoformat() if trip.start_time else None,
        "end_time": trip.end_time.isoformat() if trip.end_time else None,
        # Locations
        "start_address": trip.start_address,
        "start_lat": trip.start_lat,
        "start_lng": trip.start_lng,
        "end_address": trip.end_address,
        "end_lat": trip.end_lat,
        "end_lng": trip.end_lng,
        # Odometer
        "start_odometer": trip.start_odometer,
        "end_odometer": trip.end_odometer,
        # Purpose
        "purpose_category": trip.purpose_category,
        "purpose_notes": trip.purpose_notes,
        # ── Program transport ───────────────────────────────────────
        "scheduled_start_time": trip.scheduled_start_time,
        "scheduled_end_time": trip.scheduled_end_time,
        "out_of_schedule": trip.out_of_schedule or False,
        "out_of_schedule_note": trip.out_of_schedule_note,
        # Admin approval
        "approved_by_id": trip.approved_by_id,
        "approved_by_name": trip.approved_by.full_name if trip.approved_by else None,
        "approved_at": trip.approved_at.isoformat() if trip.approved_at else None,
        "rejection_note": trip.rejection_note,
        # Notes
        "notes": trip.notes,
        "created_at": trip.created_at.isoformat() if trip.created_at else None,
        # ── Computed Stats ────────────────────────────────────────
        "distance_km": stats["distance_km"],
        "duration_minutes": stats["duration_minutes"],
        "avg_speed_kmh": stats["avg_speed_kmh"],
        "max_speed_kmh": stats["max_speed_kmh"],
        "gps_points_count": stats["gps_points_count"],
        # GPS trace only on detail view
        "visited_points": stats["visited_points"] if include_gps else [],
    }


def _org_filter(q, current_admin: Admin, model):
    """Apply org isolation for local admins."""
    if not current_admin.is_super_admin and current_admin.organization_id:
        q = q.filter(model.organization_id == current_admin.organization_id)
    return q

# =============================================================================
# ENDPOINTS
# =============================================================================

@router.get("/")
def list_trips(
    page: int = 1,
    page_size: int = 20,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vehicle_id: Optional[str] = None,
    driver_id: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Listare foi de parcurs cu filtre."""
    q = db.query(TripLog)
    q = _org_filter(q, current_admin, TripLog)

    if date_from:
        q = q.filter(TripLog.date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(TripLog.date <= date.fromisoformat(date_to))
    if vehicle_id:
        q = q.filter(TripLog.vehicle_id == vehicle_id)
    if driver_id:
        q = q.filter(TripLog.driver_id == driver_id)
    if status:
        q = q.filter(TripLog.status == status)

    total = q.count()
    trips = q.order_by(desc(TripLog.date), desc(TripLog.start_time)) \
             .offset((page - 1) * page_size).limit(page_size).all()

    return {
        "trips": [trip_to_dict(t) for t in trips],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/stats")
def get_transport_stats(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Statistici globale transport: km totali, drumuri, viteză medie, top șoferi."""
    q = db.query(TripLog)
    q = _org_filter(q, current_admin, TripLog)

    if date_from:
        q = q.filter(TripLog.date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(TripLog.date <= date.fromisoformat(date_to))

    trips = q.all()

    total_trips = len(trips)
    completed   = [t for t in trips if t.status in ("completed", "approved")]
    in_progress = [t for t in trips if t.status == "in_progress"]
    pending_apr = [t for t in trips if t.status == "completed"]

    # KM totali — din odometru dacă disponibil, altfel GPS
    total_km = 0.0
    total_duration_min = 0.0
    all_speeds = []

    for t in completed:
        stats = compute_trip_stats(t)
        if stats["distance_km"]:
            total_km += stats["distance_km"]
        if stats["duration_minutes"]:
            total_duration_min += stats["duration_minutes"]
        if stats["avg_speed_kmh"]:
            all_speeds.append(stats["avg_speed_kmh"])

    avg_speed = round(sum(all_speeds) / len(all_speeds), 1) if all_speeds else None

    # Top 5 șoferi după km
    driver_km: dict = {}
    for t in completed:
        if not t.driver_id:
            continue
        stats = compute_trip_stats(t)
        km = stats["distance_km"] or 0
        name = t.driver.full_name if t.driver else "N/A"
        if t.driver_id not in driver_km:
            driver_km[t.driver_id] = {"name": name, "km": 0, "trips": 0}
        driver_km[t.driver_id]["km"] += km
        driver_km[t.driver_id]["trips"] += 1

    top_drivers = sorted(driver_km.values(), key=lambda x: x["km"], reverse=True)[:5]
    for d in top_drivers:
        d["km"] = round(d["km"], 1)

    # Top 5 vehicule după km
    vehicle_km: dict = {}
    for t in completed:
        if not t.vehicle_id:
            continue
        stats = compute_trip_stats(t)
        km = stats["distance_km"] or 0
        name = t.vehicle.name if t.vehicle else "N/A"
        plate = t.vehicle.plate_number if t.vehicle else ""
        if t.vehicle_id not in vehicle_km:
            vehicle_km[t.vehicle_id] = {"name": name, "plate": plate, "km": 0, "trips": 0}
        vehicle_km[t.vehicle_id]["km"] += km
        vehicle_km[t.vehicle_id]["trips"] += 1

    top_vehicles = sorted(vehicle_km.values(), key=lambda x: x["km"], reverse=True)[:5]
    for v in top_vehicles:
        v["km"] = round(v["km"], 1)

    return {
        "total_trips": total_trips,
        "completed_trips": len(completed),
        "in_progress_trips": len(in_progress),
        "pending_approval": len(pending_apr),
        "total_km": round(total_km, 1),
        "total_duration_hours": round(total_duration_min / 60, 1),
        "avg_speed_kmh": avg_speed,
        "top_drivers": top_drivers,
        "top_vehicles": top_vehicles,
    }


@router.get("/{trip_id}")
def get_trip(
    trip_id: str,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Detaliu complet foaie de parcurs cu GPS trace, viteză, km, durată."""
    trip = db.query(TripLog).filter(TripLog.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Foaia de parcurs nu a fost găsită.")
    if not current_admin.is_super_admin and trip.organization_id != current_admin.organization_id:
        raise HTTPException(status_code=403, detail="Acces interzis.")
    return trip_to_dict(trip, include_gps=True)


@router.post("/", status_code=201)
def create_trip(
    payload: TripCreate,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Creă o foaie de parcurs. Validează programul de transport al organizației."""
    vehicle = db.query(Vehicle).filter(Vehicle.id == payload.vehicle_id).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehiculul nu a fost găsit.")

    org_id = current_admin.organization_id or vehicle.organization_id
    org = db.query(Organization).filter(Organization.id == org_id).first()

    # ── Verificare program transport ────────────────────────────────
    from app.timezone import get_local_now
    now_local = get_local_now()
    sched = _check_schedule(org, now_local)

    if not sched["allowed"]:
        raise HTTPException(
            status_code=403,
            detail=f"⛔ În afara programului de transport: {sched['note']}"
        )

    trip_date = date.fromisoformat(payload.date) if payload.date else date.today()

    trip = TripLog(
        id=str(uuid.uuid4()),
        organization_id=org_id,
        vehicle_id=payload.vehicle_id,
        driver_id=payload.driver_id,
        site_id=payload.site_id,
        date=trip_date,
        status="in_progress",
        start_time=datetime.utcnow(),
        purpose_category=payload.purpose_category,
        purpose_notes=payload.purpose_notes,
        start_odometer=payload.start_odometer,
        start_address=payload.start_address,
        start_lat=payload.start_lat,
        start_lng=payload.start_lng,
        notes=payload.notes,
        created_by_admin_id=current_admin.id,
        # Snapshot program
        scheduled_start_time=sched.get("start_time"),
        scheduled_end_time=sched.get("end_time"),
        out_of_schedule=sched["out_of_schedule"],
        out_of_schedule_note=sched["note"],
    )
    db.add(trip)
    db.commit()
    db.refresh(trip)
    return trip_to_dict(trip)


@router.put("/{trip_id}/end")
def end_trip(
    trip_id: str,
    payload: TripEnd,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Oprește un drum activ. Calculează automat km și durată."""
    trip = db.query(TripLog).filter(TripLog.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Foaia nu a fost găsită.")
    if trip.status != "in_progress":
        raise HTTPException(status_code=400, detail="Drumul nu este activ.")

    trip.end_time    = datetime.utcnow()
    trip.status      = "completed"
    trip.end_address = payload.end_address
    trip.end_lat     = payload.end_lat
    trip.end_lng     = payload.end_lng
    if payload.notes:
        trip.notes = payload.notes

    # Calculez km din odometru dacă ambele valori există
    if payload.end_odometer is not None:
        trip.end_odometer = payload.end_odometer
        if trip.start_odometer is not None:
            trip.distance_km = round(payload.end_odometer - trip.start_odometer, 2)

    # Dacă nu avem odometru, calculăm din GPS
    if trip.distance_km is None and trip.gps_points:
        gps = sorted(trip.gps_points, key=lambda p: p.timestamp)
        trip.distance_km = round(sum(
            haversine_km(gps[i].latitude, gps[i].longitude,
                         gps[i+1].latitude, gps[i+1].longitude)
            for i in range(len(gps) - 1)
        ), 2)

    db.commit()
    db.refresh(trip)
    return trip_to_dict(trip, include_gps=True)


@router.put("/{trip_id}")
def update_trip(
    trip_id: str,
    payload: TripUpdate,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Editare foaie de parcurs (corectare date manuale)."""
    trip = db.query(TripLog).filter(TripLog.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Foaia nu a fost găsită.")

    for field, val in payload.model_dump(exclude_none=True).items():
        setattr(trip, field, val)

    # Recalculez km dacă ambele odometrele sunt setate
    if trip.start_odometer is not None and trip.end_odometer is not None:
        trip.distance_km = round(trip.end_odometer - trip.start_odometer, 2)

    db.commit()
    db.refresh(trip)
    return trip_to_dict(trip)


@router.delete("/{trip_id}", status_code=204)
def delete_trip(
    trip_id: str,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Șterge o foaie de parcurs."""
    trip = db.query(TripLog).filter(TripLog.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Foaia nu a fost găsită.")
    db.delete(trip)
    db.commit()
    return None


@router.post("/{trip_id}/gps")
def add_gps_points(
    trip_id: str,
    points: List[GPSPointIn],
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Batch upload puncte GPS pentru un drum activ."""
    trip = db.query(TripLog).filter(TripLog.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Foaia nu a fost găsită.")
    if trip.status != "in_progress":
        raise HTTPException(status_code=400, detail="Nu se pot adăuga puncte GPS — drumul nu este activ.")

    for p in points:
        gps_point = TripGPSPoint(
            id=str(uuid.uuid4()),
            trip_id=trip_id,
            latitude=p.latitude,
            longitude=p.longitude,
            speed_kmh=p.speed_kmh,
            accuracy_m=p.accuracy_m,
            altitude_m=p.altitude_m,
            timestamp=datetime.fromisoformat(p.timestamp.replace("Z", "+00:00")),
        )
        db.add(gps_point)

    db.commit()
    return {"added": len(points)}


@router.put("/{trip_id}/approve")
def approve_trip(
    trip_id: str,
    payload: ApprovalPayload,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Validare sau respingere foaie de parcurs de către admin."""
    trip = db.query(TripLog).filter(TripLog.id == trip_id).first()
    if not trip:
        raise HTTPException(status_code=404, detail="Foaia nu a fost găsită.")

    if payload.action == "approve":
        trip.status        = "approved"
        trip.approved_by_id = current_admin.id
        trip.approved_at   = datetime.utcnow()
        trip.rejection_note = None
    elif payload.action == "reject":
        trip.status         = "rejected"
        trip.approved_by_id = current_admin.id
        trip.approved_at    = datetime.utcnow()
        trip.rejection_note = payload.rejection_note
    else:
        raise HTTPException(status_code=400, detail="Acțiune invalidă. Folosiți 'approve' sau 'reject'.")

    db.commit()
    db.refresh(trip)
    return trip_to_dict(trip)


@router.get("/vehicle/{vehicle_id}/calendar")
def vehicle_calendar(
    vehicle_id: str,
    year: int = Query(default=None),
    month: int = Query(default=None),
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Toate drumurile unui vehicul pe o lună — pentru vizualizare calendar."""
    from datetime import date as dt_date
    import calendar

    today = dt_date.today()
    y = year or today.year
    m = month or today.month

    _, last_day = calendar.monthrange(y, m)
    d_from = dt_date(y, m, 1)
    d_to   = dt_date(y, m, last_day)

    trips = db.query(TripLog).filter(
        TripLog.vehicle_id == vehicle_id,
        TripLog.date >= d_from,
        TripLog.date <= d_to,
    ).order_by(TripLog.date).all()

    total_km = sum((t.distance_km or 0) for t in trips if t.status in ("completed", "approved"))

    return {
        "vehicle_id": vehicle_id,
        "year": y,
        "month": m,
        "total_km_month": round(total_km, 1),
        "trips": [trip_to_dict(t) for t in trips],
    }


@router.get("/export/excel")
def export_excel(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vehicle_id: Optional[str] = None,
    driver_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Export foi de parcurs în Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = db.query(TripLog)
    q = _org_filter(q, current_admin, TripLog)
    if date_from:
        q = q.filter(TripLog.date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(TripLog.date <= date.fromisoformat(date_to))
    if vehicle_id:
        q = q.filter(TripLog.vehicle_id == vehicle_id)
    if driver_id:
        q = q.filter(TripLog.driver_id == driver_id)

    trips = q.order_by(desc(TripLog.date)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Foi de Parcurs"

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill(start_color="0f172a", end_color="1e3a5f", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    headers = [
        "Data", "Vehicul", "Nr. înmatriculare", "Șofer",
        "Plecare", "Destinație", "Scop",
        "Start (oră)", "Stop (oră)", "Durată (min)",
        "KM Start", "KM Stop", "KM Parcurși",
        "Viteză Medie (km/h)", "Viteză Max (km/h)",
        "Status", "Validat de", "Note"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    total_km_sum = 0.0
    for row_idx, t in enumerate(trips, 2):
        stats = compute_trip_stats(t)
        km = stats["distance_km"] or 0
        total_km_sum += km

        row_data = [
            str(t.date),
            t.vehicle.name if t.vehicle else "—",
            t.vehicle.plate_number if t.vehicle else "—",
            t.driver.full_name if t.driver else "—",
            t.start_address or "—",
            t.end_address or (t.site.name if t.site else "—"),
            t.purpose_category or "—",
            t.start_time.strftime("%H:%M") if t.start_time else "—",
            t.end_time.strftime("%H:%M") if t.end_time else "—",
            stats["duration_minutes"] or "—",
            t.start_odometer or "—",
            t.end_odometer or "—",
            round(km, 1) if km else "—",
            stats["avg_speed_kmh"] or "—",
            stats["max_speed_kmh"] or "—",
            t.status,
            t.approved_by.full_name if t.approved_by else "—",
            t.notes or "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    # Total row
    if trips:
        tr = len(trips) + 2
        ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=tr, column=13, value=round(total_km_sum, 1)).font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=tr, column=col).border = thin_border
            ws.cell(row=tr, column=col).fill = PatternFill(start_color="E7E6E6", fill_type="solid")

    # Column widths
    widths = {1:12, 2:25, 3:16, 4:22, 5:25, 6:25, 7:20,
              8:10, 9:10, 10:12, 11:10, 12:10, 13:12,
              14:14, 15:14, 16:14, 17:18, 18:30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"foi_parcurs_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# =============================================================================
# ENDPOINT: Program Transport (GET / PUT)
# =============================================================================

@router.get("/schedule/config")
def get_schedule_config(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Citire program transport configurat pentru organizație."""
    org_id = current_admin.organization_id
    if not org_id:
        raise HTTPException(status_code=400, detail="Super adminul trebuie să selecteze o organizație.")

    org = db.query(Organization).filter(Organization.id == org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organizația nu a fost găsită.")

    allowed_days = org.transport_allowed_days or [0, 1, 2, 3, 4]
    if isinstance(allowed_days, str):
        import json
        try:
            allowed_days = json.loads(allowed_days)
        except Exception:
            allowed_days = [0, 1, 2, 3, 4]
    
    # Ensure it's a list of ints
    allowed_days = [int(d) for d in allowed_days]
    
    return {
        "transport_start_time": org.transport_start_time or "06:00",
        "transport_end_time":   org.transport_end_time   or "20:00",
        "transport_allowed_days": allowed_days,
        "transport_allowed_day_names": [DAY_NAMES_RO[d] for d in sorted(allowed_days)],
        "transport_strict_schedule": org.transport_strict_schedule or False,
    }


@router.put("/schedule/config")
def update_schedule_config(
    payload: TransportScheduleUpdate,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Actualizare program transport pentru organizație."""
    org_id = current_admin.organization_id
    if not org_id:
        raise HTTPException(status_code=400, detail="Doar adminii locali pot configura programul organizației.")

    org = db.query(Organization).filter(Organization.id == org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organizația nu a fost găsită.")

    if payload.transport_start_time is not None:
        # Validate HH:MM format
        try:
            _parse_hhmm(payload.transport_start_time)
        except (ValueError, AttributeError):
            raise HTTPException(status_code=400, detail="Format invalid pentru ora de start (folosiți HH:MM).")
        org.transport_start_time = payload.transport_start_time

    if payload.transport_end_time is not None:
        try:
            _parse_hhmm(payload.transport_end_time)
        except (ValueError, AttributeError):
            raise HTTPException(status_code=400, detail="Format invalid pentru ora de stop (folosiți HH:MM).")
        org.transport_end_time = payload.transport_end_time

    if payload.transport_allowed_days is not None:
        invalid = [d for d in payload.transport_allowed_days if d not in range(7)]
        if invalid:
            raise HTTPException(status_code=400, detail=f"Zile invalide: {invalid}. Folosiți 0 (Luni) – 6 (Duminică).")
        org.transport_allowed_days = payload.transport_allowed_days

    if payload.transport_strict_schedule is not None:
        org.transport_strict_schedule = payload.transport_strict_schedule

    db.commit()
    db.refresh(org)

    allowed_days = org.transport_allowed_days or [0, 1, 2, 3, 4]
    return {
        "message": "Program transport actualizat cu succes.",
        "transport_start_time": org.transport_start_time,
        "transport_end_time":   org.transport_end_time,
        "transport_allowed_days": allowed_days,
        "transport_allowed_day_names": [DAY_NAMES_RO[d] for d in sorted(allowed_days)],
        "transport_strict_schedule": org.transport_strict_schedule,
    }


@router.get("/schedule/check")
def check_schedule_now(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Verifică dacă în acest moment se poate porni un drum."""
    org_id = current_admin.organization_id
    org = db.query(Organization).filter(Organization.id == org_id).first() if org_id else None

    from app.timezone import get_local_now
    now_local = get_local_now()
    sched = _check_schedule(org, now_local)

    return {
        "current_time": now_local.strftime("%H:%M"),
        "current_day": DAY_NAMES_RO[now_local.weekday()],
        "can_start_trip": sched["allowed"],
        "out_of_schedule": sched["out_of_schedule"],
        "message": sched["note"] or "Program activ. Puteți porni un drum.",
        "schedule": {
            "start_time": sched.get("start_time"),
            "end_time":   sched.get("end_time"),
            "allowed_days": sched.get("allowed_days", []),
            "allowed_day_names": [DAY_NAMES_RO[d] for d in sorted(sched.get("allowed_days") or [])],
            "strict": sched.get("strict", False),
        }
    }
