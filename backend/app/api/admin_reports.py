"""
Admin Reports: Preview & Excel export for timesheets
Works with actual schema: Timesheet → TimesheetSegment → ConstructionSite
"""
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, date
from app.timezone import get_local_now, get_local_today
from typing import Optional
import io

from app.database import get_db
from app.models import (
    Timesheet, User, ConstructionSite, TimesheetSegment,
    TimesheetLine, Activity, Role, GeofencePause, Admin
)
from app.api.admin_auth import get_current_admin

router = APIRouter()


def _build_report_data(db: Session, date_from=None, date_to=None, employee_id=None, site_id=None):
    """Build report data — optimized: bulk queries, no N+1"""
    from sqlalchemy import and_

    # ── 1. Fetch matching timesheets ──────────────────────────────────────
    query = db.query(Timesheet).filter(Timesheet.owner_type == "USER")
    if date_from:
        query = query.filter(Timesheet.date >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        query = query.filter(Timesheet.date <= datetime.strptime(date_to, "%Y-%m-%d").date())
    if employee_id:
        query = query.filter(Timesheet.owner_user_id == employee_id)

    timesheets = query.order_by(Timesheet.date.desc()).all()
    if not timesheets:
        return []

    ts_ids = [ts.id for ts in timesheets]
    user_ids = list({ts.owner_user_id for ts in timesheets})

    # ── 2. Bulk-load all related data ────────────────────────────────────
    # Users + roles
    users_list = db.query(User).filter(User.id.in_(user_ids)).all()
    user_map = {u.id: u for u in users_list}

    role_ids = list({u.role_id for u in users_list if u.role_id})
    roles_list = db.query(Role).filter(Role.id.in_(role_ids)).all()
    role_map = {r.id: r for r in roles_list}

    # Segments (all at once)
    segments_list = db.query(TimesheetSegment).filter(
        TimesheetSegment.timesheet_id.in_(ts_ids)
    ).order_by(TimesheetSegment.check_in_time.asc()).all()
    seg_by_ts = {}
    for seg in segments_list:
        seg_by_ts.setdefault(seg.timesheet_id, []).append(seg)

    # Sites
    site_ids = list({seg.site_id for seg in segments_list if seg.site_id})
    sites_list = db.query(ConstructionSite).filter(ConstructionSite.id.in_(site_ids)).all()
    site_map = {s.id: s for s in sites_list}

    # Geofence pauses
    seg_ids = [seg.id for seg in segments_list]
    pauses_list = db.query(GeofencePause).filter(GeofencePause.segment_id.in_(seg_ids)).all() if seg_ids else []
    pauses_by_seg = {}
    for gp in pauses_list:
        pauses_by_seg.setdefault(gp.segment_id, []).append(gp)

    # Timesheet lines + activities
    lines_list = db.query(TimesheetLine).filter(TimesheetLine.timesheet_id.in_(ts_ids)).all()
    act_ids = list({tl.activity_id for tl in lines_list if tl.activity_id})
    acts_list = db.query(Activity).filter(Activity.id.in_(act_ids)).all() if act_ids else []
    act_map = {a.id: a for a in acts_list}
    lines_by_ts = {}
    for tl in lines_list:
        lines_by_ts.setdefault(tl.timesheet_id, []).append(tl)

    # ── 3. Build results in memory ────────────────────────────────────────
    now = get_local_now()
    results = []

    for ts in timesheets:
        user = user_map.get(ts.owner_user_id)
        if not user:
            continue

        segments = seg_by_ts.get(ts.id, [])
        if not segments:
            continue

        first_seg = segments[0]
        last_seg = segments[-1]

        # Site filter
        if site_id and first_seg.site_id != site_id:
            continue

        seg_site = site_map.get(first_seg.site_id)
        role = role_map.get(user.role_id)

        # Calculate hours
        total_worked = 0.0
        total_break = 0.0
        for seg in segments:
            end_time = seg.check_out_time or now
            seg_hours = (end_time - seg.check_in_time).total_seconds() / 3600
            seg_break = 0.0
            if seg.break_start_time:
                break_end = seg.break_end_time or now
                seg_break = (break_end - seg.break_start_time).total_seconds() / 3600

            geo_pauses = pauses_by_seg.get(seg.id, [])
            geo_secs = sum(((gp.pause_end or now) - gp.pause_start).total_seconds() for gp in geo_pauses)
            total_worked += max(0, seg_hours - seg_break - geo_secs / 3600)
            total_break += seg_break

        # Activities
        tlines = lines_by_ts.get(ts.id, [])
        activities = []
        for tl in tlines:
            act = act_map.get(tl.activity_id)
            if act:
                activities.append(f"{act.name}: {tl.quantity_numeric or 0} {tl.unit_type or ''}")

        results.append({
            "id": str(ts.id),
            "date": ts.date.isoformat() if ts.date else None,
            "employee_id": str(user.id),
            "employee_name": user.full_name,
            "employee_code": user.employee_code,
            "role": role.name if role else "—",
            "site_name": seg_site.name if seg_site else "Necunoscut",
            "check_in": first_seg.check_in_time.strftime("%H:%M") if first_seg.check_in_time else None,
            "check_out": last_seg.check_out_time.strftime("%H:%M") if last_seg.check_out_time else "—",
            "break_minutes": round(total_break * 60, 0),
            "hours_worked": round(total_worked, 2),
            "activities": "; ".join(activities) if activities else "—"
        })

    return results


@router.get("/timesheets/preview")
async def preview_timesheets(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    site_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Preview timesheet data"""
    results = _build_report_data(db, date_from, date_to, employee_id, site_id)
    total_hours = sum(r["hours_worked"] for r in results)

    return {
        "timesheets": results,
        "total": len(results),
        "total_hours": round(total_hours, 2)
    }


@router.get("/timesheets/excel")
async def export_timesheets_excel(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    site_id: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    admin: Admin = Depends(get_current_admin)
):
    """Export timesheets to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    results = _build_report_data(db, date_from, date_to, employee_id, site_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pontaje"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0f172a", end_color="1e3a5f", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["Data", "Angajat", "Cod", "Rol", "Șantier", "Intrare", "Ieșire", "Pauză (min)", "Ore Lucrate", "Activități"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    total_hours = 0
    for row_idx, r in enumerate(results, 2):
        total_hours += r["hours_worked"]
        row_data = [
            r["date"], r["employee_name"], r["employee_code"], r["role"],
            r["site_name"], r["check_in"], r["check_out"],
            int(r["break_minutes"]), r["hours_worked"], r["activities"]
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border

    # Total row
    if results:
        tr = len(results) + 2
        ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=tr, column=9, value=round(total_hours, 2)).font = Font(bold=True)
        for col in range(1, 11):
            ws.cell(row=tr, column=col).border = border
            ws.cell(row=tr, column=col).fill = PatternFill(start_color="E7E6E6", fill_type="solid")

    # Column widths
    widths = {1: 12, 2: 22, 3: 12, 4: 15, 5: 25, 6: 8, 7: 8, 8: 10, 9: 10, 10: 40}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"pontaje_{date_from or 'all'}_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
