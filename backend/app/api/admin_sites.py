"""
Admin API endpoints for construction sites management
"""
from fastapi import APIRouter, Depends, HTTPException, status, Query
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from pydantic import BaseModel, Field
from datetime import datetime, time, date
import requests
import logging

from app.database import get_db
from app.models import ConstructionSite, Admin
from app.api.admin_auth import get_current_admin

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/admin/sites", tags=["admin-sites"])


def geocode_address(address: str, county: str = None) -> dict:
    """Geocode an address using OpenStreetMap Nominatim (free, no API key needed)"""
    try:
        query = address
        if county:
            query += f", {county}"
        query += ", Romania"
        
        response = requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": query, "format": "json", "limit": 1, "countrycodes": "ro"},
            headers={"User-Agent": "PontajDigital/1.0"},
            timeout=5
        )
        results = response.json()
        if results:
            return {
                "latitude": float(results[0]["lat"]),
                "longitude": float(results[0]["lon"])
            }
    except Exception as e:
        logger.warning(f"Geocoding failed for '{address}': {e}")
    return {}


# Pydantic schemas
class SiteCreate(BaseModel):
    organization_id: Optional[str] = None
    name: str = Field(..., min_length=2, max_length=255)
    address: Optional[str] = None
    county: Optional[str] = None
    description: Optional[str] = None
    status: str = "active"  # active, completed, suspended
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    geofence_radius: Optional[int] = Field(100, ge=10, le=5000)
    
    client_id: Optional[str] = None
    client_name: Optional[str] = Field(None, max_length=255)
    panel_count: Optional[int] = Field(None, ge=0)
    system_power_kw: Optional[float] = Field(None, ge=0)
    installation_type: Optional[str] = None  # residential, commercial, industrial
    
    # Work schedule
    work_start_time: Optional[str] = "07:00"   # HH:MM format
    work_end_time: Optional[str] = "16:00"     # HH:MM format
    lunch_break_start: Optional[str] = "12:00"
    lunch_break_end: Optional[str] = "13:00"
    max_overtime_minutes: Optional[int] = Field(120, ge=0, le=480)

    # Lucrare scurta durata
    project_type: Optional[str] = "standard"   # standard | short_term
    planned_start_date: Optional[str] = None    # YYYY-MM-DD
    planned_end_date: Optional[str] = None      # YYYY-MM-DD


class SiteTeamsUpdate(BaseModel):
    team_ids: list[str]

class SiteUpdate(BaseModel):
    name: Optional[str] = Field(None, min_length=2, max_length=255)
    address: Optional[str] = None
    county: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None  # active, completed, suspended
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    geofence_radius: Optional[int] = Field(None, ge=10, le=5000)
    
    client_id: Optional[str] = None
    client_name: Optional[str] = Field(None, max_length=255)
    panel_count: Optional[int] = Field(None, ge=0)
    system_power_kw: Optional[float] = Field(None, ge=0)
    installation_type: Optional[str] = None
    
    # Work schedule
    work_start_time: Optional[str] = None   # HH:MM format
    work_end_time: Optional[str] = None     # HH:MM format
    lunch_break_start: Optional[str] = None
    lunch_break_end: Optional[str] = None
    max_overtime_minutes: Optional[int] = Field(None, ge=0, le=480)

    # Lucrare scurta durata
    project_type: Optional[str] = None
    planned_start_date: Optional[str] = None
    planned_end_date: Optional[str] = None


class QuickWorkerAssign(BaseModel):
    """Alocare rapida muncitori la o lucrare scurta."""
    user_ids: List[str]


def site_to_dict(site, db=None) -> dict:
    """Convert a ConstructionSite ORM object to a JSON-serializable dict."""
    assigned_workers = 0
    assigned_team_ids = []
    direct_ids = set()
    team_ids_set = set()
    if db:
        from app.models import Team, TeamMember, User
        from sqlalchemy import func
        
        # Combine direct users and team users to get total unique workers
        direct_users = db.query(User.id).filter(User.site_id == site.id).all()
        direct_ids = {u[0] for u in direct_users}
        
        team_users = db.query(TeamMember.user_id).join(Team, Team.id == TeamMember.team_id).filter(Team.site_id == site.id).all()
        team_ids_set = {u[0] for u in team_users}
        
        assigned_workers = len(direct_ids.union(team_ids_set))
        
        teams_db = db.query(Team).filter(Team.site_id == site.id).all()
        assigned_team_ids = [t.id for t in teams_db]

    # Calcule durata lucrare scurta
    today = date.today()
    planned_duration_days = None
    days_elapsed = None
    days_remaining = None
    progress_pct = None
    urgency = None   # on_track | urgent | overdue | completed

    if site.project_type == "short_term" and site.planned_start_date and site.planned_end_date:
        planned_duration_days = (site.planned_end_date - site.planned_start_date).days
        if site.status == "completed":
            urgency = "completed"
        else:
            days_elapsed = max(0, (today - site.planned_start_date).days)
            days_remaining = (site.planned_end_date - today).days
            progress_pct = min(100, round(days_elapsed / planned_duration_days * 100)) if planned_duration_days > 0 else 0
            if days_remaining < 0:
                urgency = "overdue"
            elif days_remaining <= 1:
                urgency = "urgent"
            else:
                urgency = "on_track"

    return {
        "id": site.id,
        "name": site.name,
        "address": site.address,
        "county": site.county,
        "description": site.description,
        "status": site.status,
        "latitude": site.latitude,
        "longitude": site.longitude,
        "geofence_radius": site.geofence_radius,
        "created_at": site.created_at.isoformat() if site.created_at else None,
        "client_id": site.client_id,
        "client_name": site.client.name if site.client_id and hasattr(site, 'client') and site.client else site.client_name,
        "panel_count": site.panel_count,
        "system_power_kw": site.system_power_kw,
        "installation_type": site.installation_type,
        "work_start_time": (site.work_start_time[:5] if isinstance(site.work_start_time, str) else site.work_start_time.strftime('%H:%M')) if site.work_start_time else "07:00",
        "work_end_time": (site.work_end_time[:5] if isinstance(site.work_end_time, str) else site.work_end_time.strftime('%H:%M')) if site.work_end_time else "16:00",
        "lunch_break_start": (site.lunch_break_start[:5] if isinstance(site.lunch_break_start, str) else site.lunch_break_start.strftime('%H:%M')) if site.lunch_break_start else "12:00",
        "lunch_break_end": (site.lunch_break_end[:5] if isinstance(site.lunch_break_end, str) else site.lunch_break_end.strftime('%H:%M')) if site.lunch_break_end else "13:00",
        "max_overtime_minutes": site.max_overtime_minutes if site.max_overtime_minutes is not None else 120,
        "assigned_workers": assigned_workers,
        "assigned_worker_ids": list(direct_ids.union(team_ids_set)),
        "team_ids": assigned_team_ids,
        # Lucrare scurta durata
        "project_type": getattr(site, 'project_type', 'standard') or 'standard',
        "planned_start_date": site.planned_start_date.isoformat() if site.planned_start_date else None,
        "planned_end_date": site.planned_end_date.isoformat() if site.planned_end_date else None,
        "planned_duration_days": planned_duration_days,
        "days_elapsed": days_elapsed,
        "days_remaining": days_remaining,
        "progress_pct": progress_pct,
        "urgency": urgency,
    }


class SitesListResponse(BaseModel):
    sites: list
    total: int
    page: int
    page_size: int


@router.get("/", response_model=SitesListResponse)
def get_sites(
    page: int = 1,
    page_size: int = 20,
    search: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """
    Get paginated list of construction sites
    """
    query = db.query(ConstructionSite)

    # Isolate data per organization for local admins
    if not current_admin.is_super_admin and current_admin.organization_id:
        query = query.filter(ConstructionSite.organization_id == current_admin.organization_id)
    
    # Apply filters
    if search:
        query = query.filter(
            ConstructionSite.name.ilike(f"%{search}%") |
            ConstructionSite.address.ilike(f"%{search}%")
        )
    
    if status:
        query = query.filter(ConstructionSite.status == status)
    
    # Get total count
    total = query.count()
    
    # Apply pagination
    offset = (page - 1) * page_size
    sites = query.order_by(ConstructionSite.created_at.desc()).offset(offset).limit(page_size).all()
    
    return {
        "sites": [site_to_dict(site, db) for site in sites],
        "total": total,
        "page": page,
        "page_size": page_size
    }


# IMPORTANT: Stats routes MUST be before /{site_id} to avoid 'stats' being matched as a site_id
@router.get("/stats/summary")
def get_sites_stats(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """
    Get construction sites statistics
    """
    base = db.query(func.count(ConstructionSite.id))
    if not current_admin.is_super_admin and current_admin.organization_id:
        base = base.filter(ConstructionSite.organization_id == current_admin.organization_id)

    total_sites = base.filter(ConstructionSite.status != "suspended").scalar()
    active_sites = db.query(func.count(ConstructionSite.id))
    if not current_admin.is_super_admin and current_admin.organization_id:
        active_sites = active_sites.filter(ConstructionSite.organization_id == current_admin.organization_id)
    active_sites = active_sites.filter(ConstructionSite.status == "active").scalar()

    completed_sites = db.query(func.count(ConstructionSite.id))
    if not current_admin.is_super_admin and current_admin.organization_id:
        completed_sites = completed_sites.filter(ConstructionSite.organization_id == current_admin.organization_id)
    completed_sites = completed_sites.filter(ConstructionSite.status == "completed").scalar()

    suspended_sites = db.query(func.count(ConstructionSite.id))
    if not current_admin.is_super_admin and current_admin.organization_id:
        suspended_sites = suspended_sites.filter(ConstructionSite.organization_id == current_admin.organization_id)
    suspended_sites = suspended_sites.filter(ConstructionSite.status == "suspended").scalar()
    
    return {
        "total_sites": total_sites,
        "active_sites": active_sites,
        "completed_sites": completed_sites,
        "suspended_sites": suspended_sites
    }


@router.get("/stats")
def get_sites_stats_alias(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """
    Alias for /stats/summary for dashboard compatibility
    """
    return get_sites_stats(db, current_admin)


@router.get("/{site_id}")
def get_site(
    site_id: str,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """
    Get single construction site by ID
    """
    site = db.query(ConstructionSite).filter(ConstructionSite.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Construction site not found")
    
    return site_to_dict(site, db)


@router.get("/{site_id}/details")
def get_site_details(
    site_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """
    Get detailed site information including related entities (vehicles, warehouse, attendance, photos)
    """
    site = db.query(ConstructionSite).filter(ConstructionSite.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Construction site not found")
    
    from app.models import (
        VehicleSiteAssignment, Vehicle, 
        WarehouseTransaction, WarehouseItem, 
        TimesheetSegment, Timesheet, User,
        SitePhoto
    )
    from sqlalchemy.orm import joinedload
    
    # Teams and direct users assigned
    from app.models import Team, TeamMember
    teams_db = db.query(Team).filter(Team.site_id == site_id).all()
    teams_list = []
    for team in teams_db:
        members = db.query(User).join(TeamMember, TeamMember.user_id == User.id).filter(TeamMember.team_id == team.id).all()
        teams_list.append({
            "id": team.id,
            "name": team.name,
            "members": [{"id": m.id, "name": m.full_name, "role": m.role} for m in members]
        })
        
    direct_users_db = db.query(User).filter(User.site_id == site_id).all()
    direct_users_list = [{"id": m.id, "name": m.full_name, "role": m.role} for m in direct_users_db]
    
    # Vehicles assigned
    vehicles_assigned = db.query(VehicleSiteAssignment, Vehicle).join(
        Vehicle, VehicleSiteAssignment.vehicle_id == Vehicle.id
    ).filter(
        VehicleSiteAssignment.site_id == site_id,
        VehicleSiteAssignment.is_active == True
    ).all()
    
    vehicles_list = [
        {
            "id": v.id,
            "name": v.name,
            "type": v.type,
            "plate_number": v.plate_number,
            "chassis_number": v.chassis_number,
            "year": v.year,
            "assigned_at": str(va.created_at)
        }
        for va, v in vehicles_assigned
    ]
    
    # Date filtering base queries
    from datetime import datetime
    sd = datetime.fromisoformat(start_date) if start_date else None
    ed = datetime.fromisoformat(end_date) if end_date else None
    
    # Warehouse transactions
    wh_query = db.query(WarehouseTransaction, WarehouseItem, User).join(
        WarehouseItem, WarehouseTransaction.item_id == WarehouseItem.id
    ).outerjoin(
        User, WarehouseTransaction.assigned_to_user_id == User.id
    ).filter(WarehouseTransaction.site_id == site_id)
    
    if sd: wh_query = wh_query.filter(WarehouseTransaction.created_at >= sd)
    if ed: wh_query = wh_query.filter(WarehouseTransaction.created_at <= ed)
    
    wh_transactions = wh_query.order_by(WarehouseTransaction.created_at.desc()).all()
    
    warehouse_list = [
        {
            "id": t.id,
            "tx_type": t.transaction_type,
            "quantity": t.quantity,
            "created_at": str(t.created_at),
            "item_name": item.name,
            "item_sku": item.category,
            "user_name": u.full_name if u else "N/A"
        }
        for t, item, u in wh_transactions
    ]
    
    # Attendance (Timesheet Segments) — with real activities
    from app.models import TimesheetLine, Activity as ActivityModel

    ts_query = db.query(TimesheetSegment, Timesheet, User).join(
        Timesheet, TimesheetSegment.timesheet_id == Timesheet.id
    ).join(
        User, Timesheet.owner_user_id == User.id
    ).filter(TimesheetSegment.site_id == site_id)
    
    if sd: ts_query = ts_query.filter(Timesheet.date >= sd.date())
    if ed: ts_query = ts_query.filter(Timesheet.date <= ed.date())
    
    ts_segments = ts_query.order_by(TimesheetSegment.check_in_time.desc()).all()

    # Build a map: segment_id -> list of activity names
    seg_ids = [seg.id for seg, _, _ in ts_segments]
    activity_map = {}
    if seg_ids:
        lines = db.query(TimesheetLine, ActivityModel).join(
            ActivityModel, TimesheetLine.activity_id == ActivityModel.id
        ).filter(TimesheetLine.segment_id.in_(seg_ids)).all()
        for line, act in lines:
            if line.segment_id not in activity_map:
                activity_map[line.segment_id] = []
            entry = act.name
            if line.quantity_numeric:
                entry += f" ({line.quantity_numeric} {line.unit_type or ''})"
            activity_map[line.segment_id].append(entry)

    attendance_list = [
        {
            "id": seg.id,
            "date": str(ts.date),
            "user_name": u.full_name,
            "check_in": str(seg.check_in_time) if seg.check_in_time else None,
            "check_out": str(seg.check_out_time) if seg.check_out_time else None,
            "activities": activity_map.get(seg.id, []),
        }
        for seg, ts, u in ts_segments
    ]
    
    # Photos
    ph_query = db.query(SitePhoto, User).outerjoin(
        User, SitePhoto.uploaded_by_user_id == User.id
    ).filter(SitePhoto.site_id == site_id)
    
    if sd: ph_query = ph_query.filter(SitePhoto.created_at >= sd)
    if ed: ph_query = ph_query.filter(SitePhoto.created_at <= ed)
    
    photos = ph_query.order_by(SitePhoto.created_at.desc()).all()
    
    photo_list = [
        {
            "id": p.id,
            "photo_path": p.photo_path,
            "description": p.description,
            "created_at": str(p.created_at),
            "uploader_name": u.full_name if u else "N/A"
        }
        for p, u in photos
    ]
    
    return {
        "site": site_to_dict(site, db),
        "teams": teams_list,
        "direct_users": direct_users_list,
        "vehicles": vehicles_list,
        "warehouse_transactions": warehouse_list,
        "attendance": attendance_list,
        "photos": photo_list
    }


@router.post("/", status_code=status.HTTP_201_CREATED)
def create_site(
    site_data: SiteCreate,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """
    Create new construction site
    """
    # Check if site with same name exists
    existing_site = db.query(ConstructionSite).filter(ConstructionSite.name == site_data.name).first()
    if existing_site:
        raise HTTPException(status_code=400, detail="Site with this name already exists")
    
    # Auto-geocode if address provided but no coordinates
    lat = site_data.latitude
    lng = site_data.longitude
    if site_data.address and not (lat and lng):
        geo = geocode_address(site_data.address, site_data.county)
        lat = geo.get("latitude")
        lng = geo.get("longitude")
    
    org_id = site_data.organization_id or current_admin.organization_id
    if not org_id:
        from app.models import Organization
        org = db.query(Organization).first()
        if org:
            org_id = org.id
        else:
            raise HTTPException(status_code=400, detail="No organization available in the system")

    new_site = ConstructionSite(
        organization_id=org_id,
        name=site_data.name,
        address=site_data.address,
        county=site_data.county,
        description=site_data.description,
        status=site_data.status,
        latitude=lat,
        longitude=lng,
        geofence_radius=site_data.geofence_radius or 100,
        client_id=site_data.client_id,
        client_name=site_data.client_name,
        panel_count=site_data.panel_count,
        system_power_kw=site_data.system_power_kw,
        installation_type=site_data.installation_type,
        work_start_time=time.fromisoformat(site_data.work_start_time) if site_data.work_start_time else time(7, 0),
        work_end_time=time.fromisoformat(site_data.work_end_time) if site_data.work_end_time else time(16, 0),
        lunch_break_start=time.fromisoformat(site_data.lunch_break_start) if site_data.lunch_break_start else time(12, 0),
        lunch_break_end=time.fromisoformat(site_data.lunch_break_end) if site_data.lunch_break_end else time(13, 0),
        max_overtime_minutes=site_data.max_overtime_minutes if site_data.max_overtime_minutes is not None else 120,
        project_type=site_data.project_type or "standard",
        planned_start_date=date.fromisoformat(site_data.planned_start_date) if site_data.planned_start_date else None,
        planned_end_date=date.fromisoformat(site_data.planned_end_date) if site_data.planned_end_date else None,
    )
    
    db.add(new_site)
    db.commit()
    db.refresh(new_site)
    
    return site_to_dict(new_site, db)


@router.put("/{site_id}")
def update_site(
    site_id: str,
    site_data: SiteUpdate,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """
    Update construction site
    """
    site = db.query(ConstructionSite).filter(ConstructionSite.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Construction site not found")
    
    # Update fields
    if site_data.name is not None:
        # Check if new name conflicts with another site
        existing = db.query(ConstructionSite).filter(
            ConstructionSite.name == site_data.name,
            ConstructionSite.id != site_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Site with this name already exists")
        site.name = site_data.name
    
    if site_data.address is not None:
        site.address = site_data.address
    
    if site_data.county is not None:
        site.county = site_data.county
    
    if site_data.description is not None:
        site.description = site_data.description
    
    if site_data.status is not None:
        site.status = site_data.status
    
    # Update geo fields
    if site_data.latitude is not None:
        site.latitude = site_data.latitude
    if site_data.longitude is not None:
        site.longitude = site_data.longitude
    if site_data.geofence_radius is not None:
        site.geofence_radius = site_data.geofence_radius
    
    # Auto-geocode if address changed but no new coordinates provided
    if site_data.address is not None and site_data.latitude is None:
        geo = geocode_address(site_data.address, site_data.county or site.county)
        if geo:
            site.latitude = geo.get("latitude")
            site.longitude = geo.get("longitude")
    
    # Update solar panel fields
    if site_data.client_id is not None:
        site.client_id = site_data.client_id
    if site_data.client_name is not None:
        site.client_name = site_data.client_name
    
    if site_data.panel_count is not None:
        site.panel_count = site_data.panel_count
    
    if site_data.system_power_kw is not None:
        site.system_power_kw = site_data.system_power_kw
    
    if site_data.installation_type is not None:
        site.installation_type = site_data.installation_type
    
    # Update schedule fields
    if site_data.work_start_time is not None:
        site.work_start_time = time.fromisoformat(site_data.work_start_time)
    if site_data.work_end_time is not None:
        site.work_end_time = time.fromisoformat(site_data.work_end_time)
    if site_data.lunch_break_start is not None:
        site.lunch_break_start = time.fromisoformat(site_data.lunch_break_start)
    if site_data.lunch_break_end is not None:
        site.lunch_break_end = time.fromisoformat(site_data.lunch_break_end)
    if site_data.max_overtime_minutes is not None:
        site.max_overtime_minutes = site_data.max_overtime_minutes

    # Lucrare scurta durata
    if site_data.project_type is not None:
        site.project_type = site_data.project_type
    if site_data.planned_start_date is not None:
        site.planned_start_date = date.fromisoformat(site_data.planned_start_date) if site_data.planned_start_date else None
    if site_data.planned_end_date is not None:
        site.planned_end_date = date.fromisoformat(site_data.planned_end_date) if site_data.planned_end_date else None

    db.commit()
    db.refresh(site)
    
    return site_to_dict(site, db)


@router.delete("/{site_id}")
def delete_site(
    site_id: str,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """
    Delete construction site (soft delete by setting status to suspended)
    """
    site = db.query(ConstructionSite).filter(ConstructionSite.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Construction site not found")
    
    # Hybrid Delete Logic
    from app.models import TimesheetSegment, Team, User
    
    has_timesheets = db.query(TimesheetSegment).filter(TimesheetSegment.site_id == site_id).first()
    
    if has_timesheets:
        # Soft delete if there are timesheets to preserve history
        site.status = "suspended"
        db.commit()
        return {"message": "Site suspended (cannot hard delete due to existing timesheets)"}
    else:
        # Hard delete: clear simple foreign keys first
        db.query(Team).filter(Team.site_id == site_id).update({"site_id": None}, synchronize_session=False)
        db.query(User).filter(User.site_id == site_id).update({"site_id": None}, synchronize_session=False)
        
        # Then safely delete the site
        db.delete(site)
        db.commit()
        return {"message": "Site permanently deleted"}
    
    return {"message": "Site deleted successfully"}


@router.get("/map-data/all")
def get_map_data(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """
    Returns all active sites with GPS coordinates and live counts
    for the Leaflet map dashboard.
    """
    from app.models import Timesheet, TimesheetSegment, VehicleSiteAssignment
    from app.timezone import get_local_today
    import sqlalchemy

    today = get_local_today()
    sites = db.query(ConstructionSite).filter(ConstructionSite.status == "active").all()
    result = []

    for site in sites:
        # Count active workers today (segments with no check_out)
        active_workers = db.execute(
            sqlalchemy.text(
                """
                SELECT COUNT(DISTINCT ts.owner_user_id)
                FROM timesheets ts
                JOIN timesheet_segments seg ON seg.timesheet_id = ts.id
                WHERE ts.date = :today
                  AND seg.site_id = :site_id
                  AND seg.check_out_time IS NULL
                """
            ),
            {"today": today.isoformat(), "site_id": site.id}
        ).scalar() or 0

        # Count vehicles assigned to this site
        vehicle_count = db.query(VehicleSiteAssignment).filter(
            VehicleSiteAssignment.site_id == site.id,
            VehicleSiteAssignment.is_active == True
        ).count()

        result.append({
            "id": site.id,
            "name": site.name,
            "address": site.address,
            "county": site.county,
            "latitude": site.latitude,
            "longitude": site.longitude,
            "geofence_radius": site.geofence_radius or 100,
            "status": site.status,
            "active_workers": int(active_workers),
            "vehicle_count": vehicle_count,
            "client_id": site.client_id,
            "panel_count": site.panel_count,
            "system_power_kw": site.system_power_kw,
            "client_name": site.client_name,
        })

    return result

@router.put("/{site_id}/teams")
def set_site_teams(
    site_id: str,
    data: SiteTeamsUpdate,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Assign teams to a site"""
    site = db.query(ConstructionSite).filter(ConstructionSite.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Construction site not found")

    from app.models import Team
    
    # Unassign all teams currently assigned to this site
    db.query(Team).filter(Team.site_id == site_id).update({"site_id": None}, synchronize_session=False)
    
    # Assign the new ones
    if data.team_ids:
        db.query(Team).filter(Team.id.in_(data.team_ids)).update({"site_id": site_id}, synchronize_session=False)
        
    db.commit()
    
    return site_to_dict(site, db)


# =============================================================================
# LUCRARI SCURTE — Endpoint-uri dedicate
# =============================================================================

@router.get("/short-term/list")
def list_short_term_sites(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """
    Listare lucrari de scurta durata cu statusul de urgenta calculat.
    """
    q = db.query(ConstructionSite).filter(ConstructionSite.project_type == "short_term")
    if not current_admin.is_super_admin and current_admin.organization_id:
        q = q.filter(ConstructionSite.organization_id == current_admin.organization_id)
    if status:
        q = q.filter(ConstructionSite.status == status)
    sites = q.order_by(ConstructionSite.planned_end_date.asc().nullslast()).all()
    return [site_to_dict(s, db) for s in sites]


@router.put("/{site_id}/workers")
def assign_workers_to_site(
    site_id: str,
    data: QuickWorkerAssign,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """
    Alocare rapida muncitori la un santier (direct, fara a edita fiecare profil).
    Seteaza User.site_id = site_id pentru toti user_ids furnizati.
    Ceilalti muncitori deja alocati la acest santier sunt dezalocati.
    """
    from app.models import User
    site = db.query(ConstructionSite).filter(ConstructionSite.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Santierul nu a fost gasit.")

    # Dezalocare muncitori care erau pe acest santier dar nu mai sunt in lista
    db.query(User).filter(
        User.site_id == site_id,
        User.id.notin_(data.user_ids)
    ).update({"site_id": None}, synchronize_session=False)

    # Alocare muncitori noi
    if data.user_ids:
        db.query(User).filter(User.id.in_(data.user_ids)).update(
            {"site_id": site_id}, synchronize_session=False
        )

    db.commit()
    return site_to_dict(site, db)


@router.get("/{site_id}/final-report")
def get_final_report(
    site_id: str,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """
    Raport final pentru o lucrare: muncitori, ore lucrate, materiale consumate, vehicule.
    Foloseste datele deja existente din pontaj, magazie si transport.
    """
    from app.models import (
        User, TimesheetSegment, Timesheet,
        WarehouseTransaction, WarehouseItem,
        TripLog, VehicleSiteAssignment, Vehicle
    )

    site = db.query(ConstructionSite).filter(ConstructionSite.id == site_id).first()
    if not site:
        raise HTTPException(status_code=404, detail="Santierul nu a fost gasit.")

    # ── Muncitori alocati ──────────────────────────────────────────────────
    direct_workers = db.query(User).filter(User.site_id == site_id).all()
    worker_ids = [w.id for w in direct_workers]

    # ── Ore lucrate per muncitor ──────────────────────────────────────────
    from sqlalchemy import text
    worker_hours = []
    total_hours = 0.0

    for w in direct_workers:
        segs = db.query(TimesheetSegment, Timesheet).join(
            Timesheet, TimesheetSegment.timesheet_id == Timesheet.id
        ).filter(
            Timesheet.owner_user_id == w.id,
            TimesheetSegment.site_id == site_id,
        ).all()
        
        hours = 0.0
        days_set = set()
        for seg, ts in segs:
            if seg.check_in_time and seg.check_out_time:
                delta = (seg.check_out_time - seg.check_in_time).total_seconds() / 3600
                hours += delta
                days_set.add(ts.date)

        total_hours += hours
        worker_hours.append({
            "user_id": w.id,
            "name": w.full_name,
            "employee_code": w.employee_code,
            "hours_worked": round(hours, 2),
            "days_present": len(days_set),
        })

    # ── Materiale consumate ───────────────────────────────────────────────
    wh_rows = db.query(WarehouseTransaction, WarehouseItem).join(
        WarehouseItem, WarehouseTransaction.item_id == WarehouseItem.id
    ).filter(
        WarehouseTransaction.site_id == site_id,
        WarehouseTransaction.transaction_type == "out",
    ).all()

    materials_used: dict = {}
    for tx, item in wh_rows:
        key = item.id
        if key not in materials_used:
            materials_used[key] = {
                "item_id": item.id,
                "name": item.name,
                "category": item.category,
                "unit": item.unit,
                "total_quantity": 0.0,
            }
        materials_used[key]["total_quantity"] += tx.quantity or 0

    # ── Vehicule folosite (din trip_logs) ─────────────────────────────────
    trip_logs = db.query(TripLog).filter(
        TripLog.site_id == site_id,
        TripLog.status.in_(["completed", "approved"]),
    ).all()

    vehicles_used: dict = {}
    total_km_trips = 0.0
    for t in trip_logs:
        if t.vehicle_id:
            vid = t.vehicle_id
            if vid not in vehicles_used:
                v = t.vehicle
                vehicles_used[vid] = {
                    "name": v.name if v else "—",
                    "plate": v.plate_number if v else "—",
                    "trips": 0,
                    "km": 0.0,
                }
            vehicles_used[vid]["trips"] += 1
            vehicles_used[vid]["km"] += t.distance_km or 0
            total_km_trips += t.distance_km or 0

    for v in vehicles_used.values():
        v["km"] = round(v["km"], 1)

    return {
        "site": site_to_dict(site, db),
        "summary": {
            "total_workers": len(direct_workers),
            "total_hours": round(total_hours, 2),
            "total_material_types": len(materials_used),
            "total_vehicle_trips": len(trip_logs),
            "total_km_trips": round(total_km_trips, 1),
        },
        "workers": sorted(worker_hours, key=lambda x: x["hours_worked"], reverse=True),
        "materials": list(materials_used.values()),
        "vehicles": list(vehicles_used.values()),
    }


@router.get("/{site_id}/final-report/excel")
def export_final_report_excel(
    site_id: str,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin),
):
    """Export raport final lucrare scurta in Excel."""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    report = get_final_report(site_id, db, current_admin)
    site_info = report["site"]
    summary = report["summary"]

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0f172a", end_color="1e3a5f", fill_type="solid")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    # ── Sheet 1: Sumar ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Sumar"
    ws1.append(["Lucrare", site_info["name"]])
    ws1.append(["Adresa", site_info["address"] or "—"])
    ws1.append(["Client", site_info["client_name"] or "—"])
    ws1.append(["Data start", str(site_info["planned_start_date"] or "—")])
    ws1.append(["Termen", str(site_info["planned_end_date"] or "—")])
    ws1.append(["Status", site_info["status"]])
    ws1.append([])
    ws1.append(["Muncitori alocati", summary["total_workers"]])
    ws1.append(["Total ore lucrate", summary["total_hours"]])
    ws1.append(["Tipuri materiale", summary["total_material_types"]])
    ws1.append(["Drumuri vehicule", summary["total_vehicle_trips"]])
    ws1.append(["KM vehicule", summary["total_km_trips"]])
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
        for cell in row:
            cell.border = thin
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 35

    # ── Sheet 2: Muncitori ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Muncitori")
    headers2 = ["Cod", "Nume", "Ore Lucrate", "Zile Prezenti"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
    for i, w in enumerate(report["workers"], 2):
        ws2.cell(i, 1, w["employee_code"]).border = thin
        ws2.cell(i, 2, w["name"]).border = thin
        ws2.cell(i, 3, w["hours_worked"]).border = thin
        ws2.cell(i, 4, w["days_present"]).border = thin
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14

    # ── Sheet 3: Materiale ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Materiale")
    headers3 = ["Denumire", "Categorie", "Cantitate", "Unitate"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
    for i, m in enumerate(report["materials"], 2):
        ws3.cell(i, 1, m["name"]).border = thin
        ws3.cell(i, 2, m["category"] or "—").border = thin
        ws3.cell(i, 3, round(m["total_quantity"], 2)).border = thin
        ws3.cell(i, 4, m["unit"] or "—").border = thin
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 10

    # ── Sheet 4: Vehicule ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Vehicule")
    headers4 = ["Vehicul", "Nr. Inmatriculare", "Nr. Drumuri", "KM Parcursi"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
    for i, v in enumerate(report["vehicles"], 2):
        ws4.cell(i, 1, v["name"]).border = thin
        ws4.cell(i, 2, v["plate"]).border = thin
        ws4.cell(i, 3, v["trips"]).border = thin
        ws4.cell(i, 4, v["km"]).border = thin
    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 12
    ws4.column_dimensions["D"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = site_info["name"].replace(" ", "_")[:40]
    fname = f"raport_final_{safe_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
