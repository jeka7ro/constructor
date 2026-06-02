"""
Admin API endpoints for user management
Includes: CRUD, ID card upload with OCR (easyocr), Excel import/export, avatar extraction
"""
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func
from typing import List, Optional
from pydantic import BaseModel, Field, field_validator
from datetime import datetime, date
from app.timezone import get_local_now, get_local_today
import hashlib
import os
import uuid
import re
import io

from app.database import get_db
from app.models import User, Role, Admin, EmployeeDocument
from app.api.admin_auth import get_current_admin
from app.storage import upload_file, delete_file, get_content_type

router = APIRouter(prefix="/admin/users", tags=["admin-users"])

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads", "id_cards")
AVATAR_DIR = os.path.join(BASE_DIR, "uploads", "avatars")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(AVATAR_DIR, exist_ok=True)


# =================== PYDANTIC SCHEMAS ===================

class UserCreate(BaseModel):
    employee_code: str = Field(..., min_length=3, max_length=20)
    last_name: str = Field(..., min_length=1, max_length=100)
    first_name: str = Field(..., min_length=1, max_length=100)
    role_id: str
    pin: str = Field(..., min_length=4, max_length=6)
    is_active: bool = True
    password: Optional[str] = None
    cnp: Optional[str] = Field(None, max_length=13)
    birth_place: Optional[str] = None
    id_card_series: Optional[str] = None
    phone: Optional[str] = Field(None, max_length=20)
    email: Optional[str] = None
    address: Optional[str] = None
    site_id: Optional[str] = None
    birth_date: Optional[str] = None
    birth_date: Optional[str] = None


class UserUpdate(BaseModel):
    employee_code: Optional[str] = Field(None, max_length=20)
    last_name: Optional[str] = Field(None, max_length=100)
    first_name: Optional[str] = Field(None, max_length=100)
    full_name: Optional[str] = Field(None, max_length=200)
    role_id: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None
    birth_date: Optional[str] = None
    cnp: Optional[str] = Field(None, max_length=13)
    birth_place: Optional[str] = None
    id_card_series: Optional[str] = None
    phone: Optional[str] = Field(None, max_length=20)
    email: Optional[str] = None
    address: Optional[str] = None
    site_id: Optional[str] = None
    hourly_rate: Optional[float] = None  # Tarif orar — confidential, admin-only

    @field_validator('employee_code', 'last_name', 'first_name', 'full_name', 'cnp', 'phone', mode='before')
    @classmethod
    def empty_str_to_none(cls, v):
        if v == "":
            return None
        return v


class UserPinReset(BaseModel):
    new_pin: str = Field(..., min_length=4, max_length=6)


class UserResponse(BaseModel):
    id: str
    employee_code: str
    full_name: str
    last_name: Optional[str] = None
    first_name: Optional[str] = None
    role_id: str
    role_name: str
    is_active: bool
    created_at: datetime
    last_login: Optional[datetime] = None
    birth_date: Optional[str] = None
    cnp: Optional[str] = None
    birth_place: Optional[str] = None
    id_card_series: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    avatar_path: Optional[str] = None
    id_card_path: Optional[str] = None
    site_id: Optional[str] = None
    site_name: Optional[str] = None
    contract_path: Optional[str] = None
    hourly_rate: Optional[float] = None  # Tarif orar — confidential

    class Config:
        from_attributes = True


class UsersListResponse(BaseModel):
    users: List[UserResponse]
    total: int
    page: int
    page_size: int


# =================== HELPER FUNCTIONS ===================

def hash_pin(pin: str) -> str:
    return hashlib.sha256(pin.encode()).hexdigest()


def split_full_name(full_name: str):
    """Split full_name into last_name and first_name"""
    parts = full_name.strip().split(' ', 1)
    if len(parts) == 2:
        return parts[0], parts[1]
    return full_name, ''


def build_user_response(user, role_name=None):
    """Build UserResponse from a User model instance"""
    rname = role_name or (user.role.name if user.role else 'N/A')
    last_name, first_name = split_full_name(user.full_name)
    return UserResponse(
        id=user.id,
        employee_code=user.employee_code,
        full_name=user.full_name,
        last_name=last_name,
        first_name=first_name,
        role_id=user.role_id,
        role_name=rname,
        is_active=user.is_active,
        created_at=user.created_at,
        last_login=getattr(user, 'last_login', None),
        birth_date=str(user.birth_date) if user.birth_date else None,
        cnp=user.cnp,
        birth_place=getattr(user, 'birth_place', None),
        id_card_series=getattr(user, 'id_card_series', None),
        phone=user.phone,
        email=user.email,
        address=user.address,
        avatar_path=user.avatar_path,
        id_card_path=user.id_card_path,
        site_id=getattr(user, 'site_id', None),
        site_name=user.site.name if getattr(user, 'site', None) else None,
        contract_path=getattr(user, 'contract_path', None),
        hourly_rate=float(user.hourly_rate) if user.hourly_rate is not None else None,
    )


def extract_id_card_data(image_path: str, raw_text: str = None) -> dict:
    """
    Extract data from Romanian ID card (Carte de Identitate) using EasyOCR.
    Extracts: Nume, Prenume, CNP, Data nașterii, Loc naștere, Serie+Număr, Domiciliu.
    Also extracts avatar photo from the ID card.
    """
    result = {
        "last_name": None,
        "first_name": None,
        "cnp": None,
        "birth_date": None,
        "birth_place": None,
        "id_card_series": None,
        "address": None,
        "avatar_path": None,
        "raw_text": None,
        "success": False,
        "message": ""
    }

    try:
        from PIL import Image
        import os
        img = None
        try:
            ext = os.path.splitext(image_path)[1].lower()
            if ext == '.pdf':
                import fitz  # PyMuPDF
                doc = fitz.open(image_path)
                page = doc.load_page(0)  # first page
                pix = page.get_pixmap(dpi=300) # render at 300 dpi for good OCR/face extraction
                mode = "RGBA" if pix.alpha else "RGB"
                img = Image.frombytes(mode, [pix.width, pix.height], pix.samples)
                if img.mode != 'RGB':
                    img = img.convert('RGB')
            else:
                img = Image.open(image_path)
        except Exception as e:
            print(f"Skipping avatar extraction (not an image or unsupported format): {e}")

        # ===== Extract avatar using OpenCV Face Detection and OCR Layout =====
        if img:
            try:
                import cv2
                import numpy as np
                from PIL import ImageOps
                
                # Apply EXIF rotation to ensure correct pixel alignment
                img = ImageOps.exif_transpose(img)
                
                open_cv_image = np.array(img.convert('RGB'))
                open_cv_image = open_cv_image[:, :, ::-1].copy() 
                
                cascade_path = cv2.data.haarcascades + 'haarcascade_frontalface_default.xml'
                face_cascade = cv2.CascadeClassifier(cascade_path)
                
                rotations = [
                    (0, None, img),
                    (90, cv2.ROTATE_90_CLOCKWISE, img.transpose(Image.ROTATE_270)),
                    (180, cv2.ROTATE_180, img.transpose(Image.ROTATE_180)),
                    (270, cv2.ROTATE_90_COUNTERCLOCKWISE, img.transpose(Image.ROTATE_90))
                ]
                
                face_crop = None
                
                for angle, cv_rot, pil_rot in rotations:
                    rotated_cv = open_cv_image if angle == 0 else cv2.rotate(open_cv_image, cv_rot)
                    gray = cv2.cvtColor(rotated_cv, cv2.COLOR_BGR2GRAY)
                    
                    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(30, 30))
                    
                    if len(faces) > 0:
                        largest_face = max(faces, key=lambda rect: rect[2] * rect[3])
                        x, y, w, h = largest_face
                        margin_x = int(w * 0.25)
                        margin_y = int(h * 0.35)
                        img_h, img_w = rotated_cv.shape[:2]
                        
                        x1 = max(0, x - margin_x)
                        y1 = max(0, y - margin_y)
                        x2 = min(img_w, x + w + margin_x)
                        y2 = min(img_h, y + h + int(margin_y * 1.5))
                        
                        face_crop = pil_rot.crop((x1, y1, x2, y2))
                        break
                
                # === FALLBACK: OCR Layout Analysis ===
                if not face_crop:
                    try:
                        import easyocr
                        reader = easyocr.Reader(['ro', 'en'], gpu=False, verbose=False)
                        ocr_results = reader.readtext(open_cv_image)
                        
                        # Filter out MRZ
                        valid_text = [r for r in ocr_results if '<' not in r[1]]
                        
                        if len(valid_text) > 3:
                            all_x = []
                            all_y = []
                            dx_sum = 0
                            dy_sum = 0
                            for (bbox, text, prob) in valid_text:
                                pt0, pt1, pt2, pt3 = bbox
                                all_x.extend([pt0[0], pt1[0], pt2[0], pt3[0]])
                                all_y.extend([pt0[1], pt1[1], pt2[1], pt3[1]])
                                dx_sum += (pt1[0] - pt0[0])
                                dy_sum += (pt1[1] - pt0[1])
                            
                            all_x.sort()
                            all_y.sort()
                            
                            TX_min = all_x[int(len(all_x) * 0.20)]
                            TX_max = all_x[int(len(all_x) * 0.80)]
                            TY_min = all_y[int(len(all_y) * 0.10)]
                            TY_max = all_y[int(len(all_y) * 0.90)]
                            
                            if abs(dx_sum) > abs(dy_sum):
                                fh = TY_max - TY_min
                                fw = fh * 0.75
                                y_start = TY_min + fh * 0.05
                                y_end = TY_max - fh * 0.05
                                if dx_sum > 0:
                                    face_crop = img.crop((max(0, TX_min - fw), y_start, TX_min, y_end))
                                else:
                                    face_crop = img.crop((TX_max, y_start, min(img.width, TX_max + fw), y_end))
                            else:
                                fw = TX_max - TX_min
                                fh = fw * 0.75
                                x_start = TX_min + fw * 0.05
                                x_end = TX_max - fw * 0.05
                                if dy_sum > 0:
                                    face_crop = img.crop((x_start, max(0, TY_min - fh), x_end, TY_min))
                                else:
                                    face_crop = img.crop((x_start, TY_max, x_end, min(img.height, TY_max + fh)))
                    except Exception as fallback_e:
                        print("EasyOCR fallback skipped:", fallback_e)

            except Exception as e:
                print(f"Extraction failed: {e}")
                
            # === FINAL FALLBACK (Outside OpenCV try/except so it runs even if OpenCV crashes) ===
            try:
                if not face_crop and img:
                    w_img, h_img = img.size
                    face_crop = img.crop((int(w_img * 0.02), int(h_img * 0.15), int(w_img * 0.35), int(h_img * 0.85)))
                    
                if face_crop:
                    avatar_filename = f"avatar_{uuid.uuid4().hex[:8]}.jpg"
                    avatar_buf = io.BytesIO()
                    face_crop.save(avatar_buf, "JPEG", quality=90)
                    avatar_url = upload_file(avatar_buf.getvalue(), f"avatars/{avatar_filename}", "image/jpeg")
                    result["avatar_path"] = avatar_url
            except Exception as fallback_e:
                print(f"Final fallback failed: {fallback_e}")
            
    # ===== Try OCR if easyocr is available =====
        full_text = ''
        if raw_text:
            full_text = raw_text.strip()
            result["raw_text"] = full_text
        else:
            try:
                import easyocr
                import numpy as np

                reader = easyocr.Reader(['ro', 'en'], gpu=False)
                img_np = np.array(img)

                # Run OCR
                ocr_results = reader.readtext(img_np, detail=1)
                
                # Extract all text
                texts = []
                for (bbox, text, prob) in ocr_results:
                    texts.append(text.strip())
                
                full_text = '\n'.join(texts)
                result["raw_text"] = full_text
            except ImportError:
                pass
        
        # Recreate list of strings from full_text for local logic 
        texts = full_text.split('\n')
        
        # ===== Extract CNP (13 digits starting with 1,2,5,6,8) =====
        # Remove MRZ lines first because they contain expiration dates + digits that can accidentally form a valid 13-digit CNP
        non_mrz_text = '\n'.join([line for line in texts if '<' not in line])
        cnp_search_text = non_mrz_text.replace(' ', '').replace('O', '0').replace('o', '0').replace('I', '1').replace('l', '1')
        
        cnp_match = re.search(r'\b([12568]\d{12})\b', cnp_search_text)
        if not cnp_match:
            # Fallback: look for 13 digits anywhere in the non-MRZ text
            m = re.search(r'([12568]\d{12})', cnp_search_text)
            if m:
                cnp_match = m
        
        if cnp_match:
            result["cnp"] = cnp_match.group(1)
            cnp = result["cnp"]
            century = '19' if cnp[0] in '12' else '20' if cnp[0] in '56' else '19'
            year = century + cnp[1:3]
            month = cnp[3:5]
            day = cnp[5:7]
            try:
                # Robust birth date generation
                # Only accept valid dates, silently ignore if bad OCR
                from datetime import datetime as dt
                dt.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
                result["birth_date"] = f"{year}-{month}-{day}"
            except Exception:
                pass
        
        # ===== Extract Serie + Număr (pattern: XX 123456) =====
        for t in texts:
            series_match = re.search(r'\b([A-Z]{2})\s*(\d{6})\b', t)
            if series_match:
                result["id_card_series"] = f"{series_match.group(1)} {series_match.group(2)}"
                break
        
        # ===== Extract names from MRZ line =====
        mrz_surname = None
        mrz_firstname = None
        for t in texts:
            cleaned = t.replace(' ', '').upper()
            mrz_match = re.search(r'IDROU[A-Z]*?([A-Z]{2,})<<([A-Z]+)', cleaned)
            if not mrz_match:
                mrz_match = re.search(r'IDROU([A-ZĂÂÎȘȚ]{2,})<<([A-ZĂÂÎȘȚ]+)', cleaned)
            if mrz_match:
                mrz_surname = mrz_match.group(1).replace('<', '').strip()
                mrz_firstname = mrz_match.group(2).replace('<', '').strip()
                break
        
        if not mrz_surname:
            for t in texts:
                cleaned = t.replace(' ', '').upper()
                if cleaned.startswith('IDROU') and '<<' in cleaned:
                    after_idrou = cleaned[5:]
                    parts = after_idrou.split('<<')
                    name_parts = [p.replace('<', '').strip() for p in parts if p.replace('<', '').strip()]
                    if len(name_parts) >= 2:
                        mrz_surname = name_parts[0]
                        mrz_firstname = name_parts[1]
                        break
                    elif len(name_parts) == 1 and len(name_parts[0]) > 3:
                        mrz_surname = name_parts[0]
                        break

        if mrz_surname:
            result["last_name"] = mrz_surname.title()
        if mrz_firstname:
            result["first_name"] = mrz_firstname.title()
        
        # ===== Fallback: Extract fields based on label detection =====
        for i, text in enumerate(texts):
            text_upper = text.upper().strip()
            
            if not result["last_name"] and ('NUME' in text_upper or 'SURNAME' in text_upper or text_upper == 'NUME/SURNAME') and 'PRENUME' not in text_upper:
                for j in range(i + 1, min(i + 3, len(texts))):
                    candidate = texts[j].strip()
                    candidate_upper = candidate.upper()
                    if candidate_upper and not any(kw in candidate_upper for kw in ['PRENUME', 'FIRST', 'NAME', 'GIVEN', '/', 'LOC', 'DOMICILIU', 'CNP', 'SEX']):
                        name_val = re.sub(r'[^A-ZĂÂÎȘȚa-zăâîșț\s-]', '', candidate).strip()
                        if name_val and len(name_val) > 1:
                            result["last_name"] = name_val.title()
                            break
            
            if not result["first_name"] and ('PRENUME' in text_upper or 'FIRST NAME' in text_upper or 'GIVEN' in text_upper):
                for j in range(i + 1, min(i + 3, len(texts))):
                    candidate = texts[j].strip()
                    candidate_upper = candidate.upper()
                    if candidate_upper and not any(kw in candidate_upper for kw in ['NUME', 'LOC', 'DOMICILIU', 'CNP', 'SEX', 'NATIONAL', 'CETĂȚENI']):
                        name_val = re.sub(r'[^A-ZĂÂÎȘȚa-zăâîșț\s-]', '', candidate).strip()
                        if name_val and len(name_val) > 1:
                            result["first_name"] = name_val.title()
                            break
            
            if 'LOC' in text_upper and ('NAȘTERE' in text_upper or 'NASTERE' in text_upper or 'BIRTH' in text_upper):
                for j in range(i + 1, min(i + 3, len(texts))):
                    candidate = texts[j].strip()
                    candidate_upper = candidate.upper()
                    if candidate_upper and not any(kw in candidate_upper for kw in ['DOMICILIU', 'CNP', 'VALID', 'SERIE']):
                        if len(candidate) > 2:
                            result["birth_place"] = candidate.title()
                            break
            
            if 'DOMICILIU' in text_upper or 'ADDRESS' in text_upper or 'DOMICILIUL' in text_upper:
                addr_parts = []
                for j in range(i + 1, min(i + 5, len(texts))):
                    candidate = texts[j].strip()
                    candidate_upper = candidate.upper()
                    if any(kw in candidate_upper for kw in ['CNP', 'VALID', 'SERIE', 'IDROU', 'EMIS', 'CHIP']):
                        break
                    if len(candidate) > 2:
                        addr_parts.append(candidate)
                if addr_parts:
                    result["address"] = ', '.join(addr_parts)
        
        result["success"] = bool(result["cnp"] or result["last_name"] or result["first_name"])
        result["message"] = "Date extrase cu succes din cartea de identitate" if result["success"] else "Nu s-au putut extrage date din imagine"
            
    except Exception as e:
        result["message"] = f"Eroare procesare poză: {str(e)}"

    return result


# =================== API ENDPOINTS ===================

@router.get("/", response_model=UsersListResponse)
def get_users(
    page: int = 1,
    page_size: int = 20,
    search: Optional[str] = None,
    role_id: Optional[str] = None,
    is_active: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get paginated list of users with optional filters"""
    query = db.query(User).join(Role)
    if search:
        query = query.filter(or_(
            User.employee_code.ilike(f"%{search}%"),
            User.full_name.ilike(f"%{search}%")
        ))
    if role_id:
        query = query.filter(User.role_id == role_id)
    if is_active is not None:
        query = query.filter(User.is_active == is_active)
    
    total = query.count()
    offset = (page - 1) * page_size
    users = query.offset(offset).limit(page_size).all()
    return UsersListResponse(users=[build_user_response(u) for u in users], total=total, page=page, page_size=page_size)


@router.get("/stats/summary")
def get_users_stats(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get user statistics"""
    total_users = db.query(func.count(User.id)).filter(User.is_active == True).scalar()
    active_users = db.query(func.count(User.id)).filter(User.is_active == True).scalar()
    inactive_users = db.query(func.count(User.id)).filter(User.is_active == False).scalar()
    users_by_role = db.query(Role.name, func.count(User.id).label('count')).join(User).group_by(Role.name).all()
    return {
        "total_users": total_users,
        "active_users": active_users,
        "inactive_users": inactive_users,
        "users_by_role": [{"role": role, "count": count} for role, count in users_by_role]
    }


@router.get("/next-code")
def get_next_employee_code(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Get next available employee code (EMP001, EMP002, etc.)"""
    # Find all existing EMP codes and get the highest number
    existing_codes = db.query(User.employee_code).filter(
        User.employee_code.ilike("EMP%")
    ).all()
    
    max_num = 0
    for (code,) in existing_codes:
        # Extract number from EMP### format
        match = re.search(r'EMP(\d+)', code, re.IGNORECASE)
        if match:
            num = int(match.group(1))
            if num > max_num:
                max_num = num
    
    next_num = max_num + 1
    next_code = f"EMP{next_num:03d}"
    
    return {"next_code": next_code}


@router.get("/export/excel")
def export_users_excel(
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Export all users to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    users = db.query(User).join(Role).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Angajați"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["Cod Angajat", "Nume", "Prenume", "Rol", "CNP", "Serie Buletin",
               "Data Nașterii", "Loc Naștere", "Telefon", "Email", "Adresă", "Status"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, user in enumerate(users, 2):
        last_name, first_name = split_full_name(user.full_name)
        row_data = [
            user.employee_code, last_name, first_name,
            user.role.name if user.role else '',
            user.cnp or '', getattr(user, 'id_card_series', '') or '',
            str(user.birth_date) if user.birth_date else '',
            getattr(user, 'birth_place', '') or '',
            user.phone or '', user.email or '', user.address or '',
            'Activ' if user.is_active else 'Inactiv'
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"angajati_{get_local_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.post("/import/excel")
async def import_users_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Import users from Excel file"""
    from openpyxl import load_workbook

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Fișierul trebuie să fie .xlsx sau .xls")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else []
    header_map = {}
    for i, h in enumerate(headers):
        if h:
            h_lower = h.lower().strip()
            if 'cod' in h_lower or 'employee' in h_lower:
                header_map['employee_code'] = i
            elif h_lower in ('nume', 'last name', 'last_name', 'surname'):
                header_map['last_name'] = i
            elif h_lower in ('prenume', 'first name', 'first_name', 'given name'):
                header_map['first_name'] = i
            elif 'rol' in h_lower or 'role' in h_lower:
                header_map['role'] = i
            elif 'cnp' in h_lower:
                header_map['cnp'] = i
            elif 'serie' in h_lower or 'buletin' in h_lower:
                header_map['id_card_series'] = i
            elif 'nașterii' in h_lower or 'nasterii' in h_lower or 'birth' in h_lower and 'loc' not in h_lower:
                header_map['birth_date'] = i
            elif 'loc' in h_lower and ('naștere' in h_lower or 'nastere' in h_lower or 'birth' in h_lower):
                header_map['birth_place'] = i
            elif 'telefon' in h_lower or 'phone' in h_lower:
                header_map['phone'] = i
            elif 'email' in h_lower:
                header_map['email'] = i
            elif 'adres' in h_lower or 'address' in h_lower:
                header_map['address'] = i

    if 'employee_code' not in header_map:
        raise HTTPException(status_code=400, detail="Coloana 'Cod Angajat' nu a fost găsită")

    roles = {r.name.lower(): r for r in db.query(Role).all()}
    default_role = db.query(Role).first()
    created, updated, errors = 0, 0, []

    def get_val(row, key):
        if key not in header_map:
            return None
        idx = header_map[key]
        if idx < len(row) and row[idx]:
            return str(row[idx]).strip()
        return None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            emp_code = get_val(row, 'employee_code')
            if not emp_code:
                continue

            last_name = get_val(row, 'last_name') or ''
            first_name = get_val(row, 'first_name') or ''
            full_name = f"{last_name} {first_name}".strip() or emp_code

            role_name = (get_val(row, 'role') or '').lower()
            role = roles.get(role_name, default_role)
            if not role:
                errors.append(f"Rândul {row_idx}: Rol invalid")
                continue

            existing = db.query(User).filter(User.employee_code == emp_code).first()
            if existing:
                existing.full_name = full_name
                for field in ['cnp', 'phone', 'email', 'address', 'id_card_series', 'birth_place']:
                    val = get_val(row, field)
                    if val:
                        setattr(existing, field, val)
                bd = get_val(row, 'birth_date')
                if bd:
                    existing.birth_date = bd
                updated += 1
            else:
                new_user = User(
                    organization_id=role.organization_id,
                    employee_code=emp_code, full_name=full_name,
                    role_id=role.id, pin_hash=hash_pin('1234'), is_active=True,
                    cnp=get_val(row, 'cnp'), phone=get_val(row, 'phone'),
                    email=get_val(row, 'email'), address=get_val(row, 'address'),
                    id_card_series=get_val(row, 'id_card_series'),
                    birth_place=get_val(row, 'birth_place'),
                )
                bd = get_val(row, 'birth_date')
                if bd:
                    new_user.birth_date = bd
                db.add(new_user)
                created += 1
        except Exception as e:
            errors.append(f"Rândul {row_idx}: {str(e)}")

    db.commit()
    return {"message": f"Import finalizat: {created} creați, {updated} actualizați", "created": created, "updated": updated, "errors": errors[:10]}


@router.get("/{user_id}", response_model=UserResponse)
def get_user(user_id: str, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return build_user_response(user)


@router.post("/", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(user_data: UserCreate, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    existing = db.query(User).filter(User.employee_code == user_data.employee_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Codul de angajat există deja")
    role = db.query(Role).filter(Role.id == user_data.role_id).first()
    if not role:
        raise HTTPException(status_code=400, detail="Rol invalid")

    # Only super admins can create users with Super Administrator role
    if role.name == 'Super Administrator' and not getattr(current_admin, 'is_super_admin', False):
        raise HTTPException(status_code=403, detail="Doar Super Administratorul poate crea conturi de Super Administrator.")

    full_name = f"{user_data.last_name} {user_data.first_name}".strip()
    
    # Convert birth_date string to date object for SQLite
    birth_date_val = None
    if user_data.birth_date:
        from datetime import datetime as dt
        try:
            birth_date_val = dt.strptime(user_data.birth_date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            birth_date_val = None
    
    new_user = User(
        organization_id=role.organization_id,
        employee_code=user_data.employee_code,
        full_name=full_name, role_id=user_data.role_id,
        pin_hash=hash_pin(user_data.pin), is_active=user_data.is_active,
        birth_date=birth_date_val, cnp=user_data.cnp,
        birth_place=user_data.birth_place, id_card_series=user_data.id_card_series,
        phone=user_data.phone, email=user_data.email, address=user_data.address
    )
    db.add(new_user)
    
    # Also create Admin record if role is an admin role
    if role.name in ADMIN_ROLE_NAMES:
        if not user_data.email:
            raise HTTPException(status_code=400, detail="Email is required for Administrator accounts")
        if not user_data.password:
            raise HTTPException(status_code=400, detail="Password is required for Administrator accounts")
        
        # Check if email is already used by an admin
        existing_admin = db.query(Admin).filter(Admin.email == user_data.email).first()
        if existing_admin:
            raise HTTPException(status_code=400, detail="Adresa de email este deja folosită de alt administrator")
            
        new_admin = Admin(
            id=str(uuid.uuid4()),
            organization_id=role.organization_id,
            email=user_data.email,
            full_name=full_name,
            password_hash=hashlib.sha256(user_data.password.encode()).hexdigest(),
            role="SUPER_ADMIN" if role.name == "Super Administrator" else "ADMIN",
            is_active=user_data.is_active,
            is_super_admin=True if role.name == "Super Administrator" else False
        )
        db.add(new_admin)

    db.commit()
    db.refresh(new_user)
    return build_user_response(new_user, role.name)


@router.put("/{user_id}", response_model=UserResponse)
def update_user(user_id: str, user_data: UserUpdate, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Update full_name from last_name/first_name
    if user_data.last_name is not None or user_data.first_name is not None:
        current_last, current_first = split_full_name(user.full_name)
        last = user_data.last_name if user_data.last_name is not None else current_last
        first = user_data.first_name if user_data.first_name is not None else current_first
        user.full_name = f"{last} {first}".strip()
    elif user_data.full_name is not None:
        user.full_name = user_data.full_name

    # Check if trying to assign an admin role — only super admin can do this
    if user_data.role_id is not None:
        role = db.query(Role).filter(Role.id == user_data.role_id).first()
        if not role:
            raise HTTPException(status_code=400, detail="Invalid role ID")
        if role.name == 'Super Administrator' and not getattr(current_admin, 'is_super_admin', False):
            raise HTTPException(status_code=403, detail="Doar Super Administratorul poate atribui roluri de Super Administrator.")
            
        # Check if the user being edited is currently a Super Administrator
        current_role = db.query(Role).filter(Role.id == user.role_id).first()
        if current_role and current_role.name == 'Super Administrator' and not getattr(current_admin, 'is_super_admin', False):
            raise HTTPException(status_code=403, detail="Nu ai permisiunea de a modifica un Super Administrator.")
            
        user.role_id = user_data.role_id

    for field in ['employee_code', 'is_active', 'cnp', 'birth_place', 'id_card_series', 'phone', 'email', 'address']:
        val = getattr(user_data, field, None)
        if val is not None:
            if field == 'employee_code':
                # Check uniqueness manually to give a nice error message
                existing = db.query(User).filter(User.employee_code == val, User.id != user_id).first()
                if existing:
                    raise HTTPException(status_code=400, detail="Codul de angajat este deja utilizat de altcineva.")
            setattr(user, field, val)

    # hourly_rate — handle separately (0.0 is a valid value)
    if user_data.hourly_rate is not None:
        user.hourly_rate = user_data.hourly_rate

    if user_data.site_id is not None:
        user.site_id = user_data.site_id if user_data.site_id != "" else None

    # Handle birth_date separately - convert string to date object for SQLite
    if user_data.birth_date is not None:
        if user_data.birth_date:
            from datetime import datetime as dt
            try:
                user.birth_date = dt.strptime(user_data.birth_date, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                user.birth_date = None
        else:
            user.birth_date = None

    # Handle Admin record update if user has an admin role
    user_role = db.query(Role).filter(Role.id == user.role_id).first()
    if user_role and user_role.name in {'Administrator', 'Super Administrator'}:
        if user_data.email and user_data.email != user.email:
            # Check if new email is taken
            existing_admin = db.query(Admin).filter(Admin.email == user_data.email, Admin.email != user.email).first()
            if existing_admin:
                raise HTTPException(status_code=400, detail="Noua adresă de email este deja folosită de alt administrator")
                
        # Try to find the existing admin record using the old email or new email
        email_to_search = user.email
        admin_record = None
        if email_to_search:
            admin_record = db.query(Admin).filter(Admin.email == email_to_search).first()
            
        if admin_record:
            admin_record.full_name = user.full_name
            admin_record.is_active = user.is_active
            admin_record.role = "SUPER_ADMIN" if user_role.name == "Super Administrator" else "ADMIN"
            admin_record.is_super_admin = True if user_role.name == "Super Administrator" else False
            if user_data.email:
                admin_record.email = user_data.email
            if user_data.password:
                admin_record.password_hash = hashlib.sha256(user_data.password.encode()).hexdigest()
        else:
            # Create the missing admin record if they have an email
            if user_data.email or user.email:
                target_email = user_data.email or user.email
                if user_data.password:
                    new_admin = Admin(
                        id=str(uuid.uuid4()),
                        organization_id=user.organization_id,
                        email=target_email,
                        full_name=user.full_name,
                        password_hash=hashlib.sha256(user_data.password.encode()).hexdigest(),
                        role="SUPER_ADMIN" if user_role.name == "Super Administrator" else "ADMIN",
                        is_active=user.is_active,
                        is_super_admin=True if user_role.name == "Super Administrator" else False
                    )
                    db.add(new_admin)

    db.commit()
    db.refresh(user)
    return build_user_response(user)


@router.post("/{user_id}/reset-pin")
def reset_user_pin(user_id: str, pin_data: UserPinReset, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.pin_hash = hash_pin(pin_data.new_pin)
    db.commit()
    return {"message": "PIN resetat cu succes"}


@router.post("/{user_id}/upload-id-card")
async def upload_id_card(user_id: str, file: UploadFile = File(...), raw_text: Optional[str] = Form(None), db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    """Upload ID card image, run OCR extraction, and extract avatar photo"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    allowed = ('.jpg', '.jpeg', '.png', '.webp', '.bmp', '.pdf')
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"Format neacceptat. Acceptăm: {', '.join(allowed)}")

    filename = f"{user_id}_id_card{ext}"
    storage_path = f"id_cards/{filename}"

    content = await file.read()
    
    # Upload to storage
    id_card_url = upload_file(content, storage_path, get_content_type(filename))
    user.id_card_path = id_card_url

    # Save temp file for OCR processing
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(content)
        temp_path = tmp.name

    # Run OCR on temp file
    ocr_result = extract_id_card_data(temp_path, raw_text)
    
    # Clean up temp file
    try:
        os.remove(temp_path)
    except:
        pass

    # Save avatar from OCR if extracted
    if ocr_result.get("avatar_path"):
        user.avatar_path = ocr_result["avatar_path"]

    # Auto-fill empty fields from OCR
    if ocr_result.get("success"):
        if not user.cnp and ocr_result.get("cnp"):
            user.cnp = ocr_result["cnp"]
        if not user.birth_date and ocr_result.get("birth_date"):
            from datetime import datetime as dt
            try:
                user.birth_date = dt.strptime(ocr_result["birth_date"], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                pass
        if not user.address and ocr_result.get("address"):
            user.address = ocr_result["address"]
        if not getattr(user, 'birth_place', None) and ocr_result.get("birth_place"):
            user.birth_place = ocr_result["birth_place"]
        if not getattr(user, 'id_card_series', None) and ocr_result.get("id_card_series"):
            user.id_card_series = ocr_result["id_card_series"]

    db.commit()

    return {
        "message": "Carte de identitate încărcată cu succes",
        "id_card_path": user.id_card_path,
        "avatar_path": user.avatar_path,
        "ocr": ocr_result
    }


@router.post("/{user_id}/upload-contract")
async def upload_contract(user_id: str, file: UploadFile = File(...), db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    """Upload employment contract (PDF/JPG) for a user"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    allowed = ('.jpg', '.jpeg', '.png', '.pdf')
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"Format neacceptat. Acceptăm: {', '.join(allowed)}")

    filename = f"{user_id}_contract{ext}"
    storage_path = f"contracts/{filename}"

    content = await file.read()
    content_type = "application/pdf" if ext == ".pdf" else get_content_type(filename)
    contract_url = upload_file(content, storage_path, content_type)
    user.contract_path = contract_url
    db.commit()

    return {
        "message": "Contract încărcat cu succes",
        "contract_path": user.contract_path
    }


@router.delete("/{user_id}/contract")
async def delete_contract(user_id: str, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    """Delete employment contract"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.contract_path = None
    db.commit()
    return {"message": "Contract șters cu succes"}


@router.post("/ocr/extract")
async def ocr_extract_only(file: UploadFile = File(...), raw_text: Optional[str] = Form(None), current_admin: Admin = Depends(get_current_admin)):
    """Extract data from ID card image without saving to a user — used for pre-filling forms"""
    allowed = ('.jpg', '.jpeg', '.png', '.webp', '.bmp', '.pdf')
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"Format neacceptat")

    temp_filename = f"temp_{uuid.uuid4().hex}{ext}"
    temp_path = os.path.join(UPLOAD_DIR, temp_filename)

    content = await file.read()
    with open(temp_path, 'wb') as f:
        f.write(content)

    try:
        ocr_result = extract_id_card_data(temp_path, raw_text)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return ocr_result


@router.delete("/{user_id}")
def delete_user(user_id: str, hard_delete: bool = False, db: Session = Depends(get_db), current_admin: Admin = Depends(get_current_admin)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Only super admin can delete Super Administrators
    if user.role and user.role.name == 'Super Administrator' and not getattr(current_admin, 'is_super_admin', False):
        raise HTTPException(status_code=403, detail="Nu ai permisiunea de a șterge un Super Administrator.")

    if hard_delete:
        db.delete(user)
        db.commit()
        return {"message": "Utilizator șters definitiv din sistem"}
    else:
        user.is_active = False
        db.commit()
        return {"message": "Utilizator arhivat cu succes (mutat în Arhivă)"}


@router.post("/{user_id}/upload-avatar")
async def upload_avatar(
    user_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    """Upload avatar photo directly for a user"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    contents = await file.read()
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
    avatar_filename = f"avatar_{uuid.uuid4().hex[:8]}.{ext}"
    content_type = get_content_type(file.filename)
    
    avatar_url = upload_file(contents, f"avatars/{avatar_filename}", content_type)
    user.avatar_path = avatar_url
    db.commit()
    
    return {"avatar_path": avatar_url}


@router.get("/{user_id}/analytics")
def get_user_analytics(
    user_id: str,
    db: Session = Depends(get_db),
    current_admin: Admin = Depends(get_current_admin)
):
    from app.models import Timesheet, TimesheetSegment, TimesheetLine, Activity, EquipmentDailyLog, AccommodationAssignment, Accommodation, ConstructionSite, WarehouseTransaction, WarehouseItem
    from sqlalchemy import func, desc
    import datetime
    from calendar import monthrange

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    today = datetime.date.today()
    first_day_of_month = today.replace(day=1)
    
    # 1. ACCOMMODATION
    accomm_assignment = db.query(AccommodationAssignment).filter(
        AccommodationAssignment.user_id == user_id
    ).order_by(desc(AccommodationAssignment.assigned_from)).first()
    
    accommodation = None
    if accomm_assignment:
        acc = db.query(Accommodation).filter(Accommodation.id == accomm_assignment.accommodation_id).first()
        if acc:
            accommodation = {
                "name": acc.name,
                "address": acc.address,
                "city": acc.address
            }

    def calculate_segments_hours(segments_list):
        total_seconds = 0
        for seg in segments_list:
            if seg.check_in_time and seg.check_out_time:
                diff = (seg.check_out_time - seg.check_in_time).total_seconds()
                break_diff = 0
                if seg.break_start_time and seg.break_end_time:
                    break_diff = (seg.break_end_time - seg.break_start_time).total_seconds()
                total_seconds += max(0, diff - break_diff)
        return total_seconds / 3600.0

    # 2. SITE STATS (Average per worker)
    site_id = user.site_id
    site_avg_hours = 0
    site_avg_fuel = 0
    
    if site_id:
        # Get all users on this site
        site_users = db.query(User).filter(User.site_id == site_id, User.is_active == True).all()
        site_user_ids = [u.id for u in site_users]
        num_users = len(site_user_ids)
        
        if num_users > 0:
            # Avg hours this month
            site_segments = db.query(TimesheetSegment).join(Timesheet).filter(
                Timesheet.owner_user_id.in_(site_user_ids),
                Timesheet.date >= first_day_of_month,
                Timesheet.date <= today
            ).all()
            site_avg_hours = calculate_segments_hours(site_segments) / num_users
            
            # Avg fuel this month
            site_equipment_fuel = db.query(func.sum(EquipmentDailyLog.refuel_liters)).filter(
                EquipmentDailyLog.operator_id.in_(site_user_ids),
                EquipmentDailyLog.date >= first_day_of_month,
                EquipmentDailyLog.date <= today
            ).scalar() or 0
            
            site_warehouse_fuel = db.query(func.sum(WarehouseTransaction.quantity)).join(WarehouseItem).filter(
                WarehouseTransaction.assigned_to_user_id.in_(site_user_ids),
                WarehouseTransaction.transaction_type == "CONSUME",
                WarehouseItem.category == "COMBUSTIBIL",
                WarehouseTransaction.date >= first_day_of_month,
                WarehouseTransaction.date <= today
            ).scalar() or 0
            
            site_avg_fuel = (float(site_equipment_fuel) + float(site_warehouse_fuel)) / num_users

    # 3. USER STATS (This month)
    user_segments = db.query(TimesheetSegment).join(Timesheet).filter(
        Timesheet.owner_user_id == user_id,
        Timesheet.date >= first_day_of_month,
        Timesheet.date <= today
    ).all()
    user_hours = calculate_segments_hours(user_segments)

    # Fuel this month (from equipment logs)
    equipment_fuel = db.query(func.sum(EquipmentDailyLog.refuel_liters)).filter(
        EquipmentDailyLog.operator_id == user_id,
        EquipmentDailyLog.date >= first_day_of_month,
        EquipmentDailyLog.date <= today
    ).scalar() or 0
    
    # Fuel this month (from warehouse consumable transactions)
    warehouse_fuel = db.query(func.sum(WarehouseTransaction.quantity)).join(WarehouseItem).filter(
        WarehouseTransaction.assigned_to_user_id == user_id,
        WarehouseTransaction.transaction_type == "CONSUME",
        WarehouseItem.category == "COMBUSTIBIL",
        WarehouseTransaction.date >= first_day_of_month,
        WarehouseTransaction.date <= today
    ).scalar() or 0
    
    user_fuel = float(equipment_fuel) + float(warehouse_fuel)

    # Fuel received — OUT-urile din magazie pentru COMBUSTIBIL
    # Sursa 1: direct pe numele angajatului
    fuel_received_direct = db.query(func.sum(WarehouseTransaction.quantity)).join(WarehouseItem).filter(
        WarehouseTransaction.assigned_to_user_id == user_id,
        WarehouseTransaction.transaction_type == "OUT",
        WarehouseItem.category == "COMBUSTIBIL",
        WarehouseTransaction.date >= first_day_of_month,
        WarehouseTransaction.date <= today
    ).scalar() or 0

    # Sursa 2: pe șantierele unde a consumat efectiv în luna asta
    consume_site_ids = [
        row[0] for row in db.query(WarehouseTransaction.site_id).join(WarehouseItem).filter(
            WarehouseTransaction.assigned_to_user_id == user_id,
            WarehouseTransaction.transaction_type == "CONSUME",
            WarehouseItem.category == "COMBUSTIBIL",
            WarehouseTransaction.date >= first_day_of_month,
            WarehouseTransaction.date <= today,
            WarehouseTransaction.site_id != None
        ).distinct().all()
    ]

    fuel_received_at_sites = 0
    if consume_site_ids:
        fuel_received_at_sites = db.query(func.sum(WarehouseTransaction.quantity)).join(WarehouseItem).filter(
            WarehouseTransaction.site_id.in_(consume_site_ids),
            WarehouseTransaction.transaction_type == "OUT",
            WarehouseItem.category == "COMBUSTIBIL",
            WarehouseTransaction.date >= first_day_of_month,
            WarehouseTransaction.date <= today
        ).scalar() or 0

    warehouse_fuel_received = max(float(fuel_received_direct), float(fuel_received_at_sites))
    
    # 4. ACTIVITIES BREAKDOWN & ANOMALIES
    lines = db.query(TimesheetLine, Activity.name).join(Activity).join(TimesheetSegment).join(Timesheet).filter(
        Timesheet.owner_user_id == user_id,
        Timesheet.date >= first_day_of_month,
        Timesheet.date <= today
    ).all()
    
    activities_dict = {}
    for line, act_name in lines:
        if act_name not in activities_dict:
            activities_dict[act_name] = {"count": 0, "quantity": 0, "unit": line.unit_type}
        activities_dict[act_name]["count"] += 1
        if line.quantity_numeric:
            activities_dict[act_name]["quantity"] += float(line.quantity_numeric)
            
    activities_breakdown = []
    for act_name, data in activities_dict.items():
        activities_breakdown.append({
            "name": act_name,
            "count": data["count"],
            "quantity": data["quantity"],
            "unit": data["unit"]
        })
    activities_breakdown.sort(key=lambda x: x["count"], reverse=True)

    # Calculate 6-month historical chart for this user vs site
    historical_chart = []
    for i in range(5, -1, -1):
        # Calculate month range
        m = today.month - i
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        
        start_dt = datetime.date(y, m, 1)
        _, last_day = monthrange(y, m)
        end_dt = datetime.date(y, m, last_day)
        
        # User hours in month
        u_segs = db.query(TimesheetSegment).join(Timesheet).filter(
            Timesheet.owner_user_id == user_id,
            Timesheet.date >= start_dt,
            Timesheet.date <= end_dt
        ).all()
        u_h = calculate_segments_hours(u_segs)
        
        # Site avg hours in month
        s_avg = 0
        if site_id:
            s_segs = db.query(TimesheetSegment).join(Timesheet).join(User, Timesheet.owner_user_id == User.id).filter(
                User.site_id == site_id,
                Timesheet.date >= start_dt,
                Timesheet.date <= end_dt
            ).all()
            s_h = calculate_segments_hours(s_segs)
            
            s_users_count = db.query(func.count(User.id)).filter(User.site_id == site_id).scalar() or 1
            s_avg = float(s_h) / s_users_count if s_users_count > 0 else 0
            
        month_name = start_dt.strftime("%b %Y")
        historical_chart.append({
            "month": month_name,
            "user_hours": round(float(u_h), 1),
            "site_avg": round(float(s_avg), 1)
        })

    return {
        "accommodation": accommodation,
        "this_month": {
            "user_hours": round(float(user_hours), 1),
            "user_fuel": round(float(user_fuel), 1),
            "user_fuel_received": round(float(warehouse_fuel_received), 1),
            "site_avg_hours": round(float(site_avg_hours), 1),
            "site_avg_fuel": round(float(site_avg_fuel), 1),
        },
        "activities_breakdown": activities_breakdown,
        "historical_chart": historical_chart,
        "anomalies": [] # For future AI logic
    }


# ─────────────────────────────────────────────────────────────────────────────
# EMPLOYEE DOCUMENTS (ACTE ADITIONALE)
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{user_id}/documents")
def get_employee_documents(user_id: str, db: Session = Depends(get_db)):
    docs = db.query(EmployeeDocument).filter(EmployeeDocument.user_id == user_id).order_by(EmployeeDocument.uploaded_at.desc()).all()
    return [{"id": d.id, "name": d.name, "file_path": f"/api{d.file_path}", "uploaded_at": d.uploaded_at} for d in docs]

@router.post("/{user_id}/documents")
async def upload_employee_document(user_id: str, name: str = Form(...), file: UploadFile = File(...), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    ext = os.path.splitext(file.filename)[1]
    filename = f"doc_{uuid.uuid4().hex[:8]}{ext}"
    upload_dir = "uploads/documents"
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, filename)
    
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
        
    new_doc = EmployeeDocument(
        user_id=user_id,
        name=name,
        file_path=f"/uploads/documents/{filename}"
    )
    db.add(new_doc)
    db.commit()
    db.refresh(new_doc)
    return {"id": new_doc.id, "name": new_doc.name, "file_path": f"/api{new_doc.file_path}", "uploaded_at": new_doc.uploaded_at}

@router.delete("/{user_id}/documents/{doc_id}")
def delete_employee_document(user_id: str, doc_id: str, db: Session = Depends(get_db)):
    doc = db.query(EmployeeDocument).filter(EmployeeDocument.id == doc_id, EmployeeDocument.user_id == user_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
        
    try:
        if os.path.exists(doc.file_path.lstrip("/")):
            os.remove(doc.file_path.lstrip("/"))
        elif os.path.exists(doc.file_path.replace("/api/", "")):
            os.remove(doc.file_path.replace("/api/", ""))
    except:
        pass
        
    db.delete(doc)
    db.commit()
    return {"status": "ok"}
